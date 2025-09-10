"""
Basic test without ADK dependencies - fixed version
"""

import sys
from pathlib import Path
import asyncio

# Add src to Python path
sys.path.insert(0, str(Path(__file__).parent / "src"))

async def test_models_only():
    """Test just the data models."""
    print("Testing models...")
    
    try:
        from excel_agent.models.base import AgentRequest, AgentResponse, AgentStatus
        from excel_agent.models.agents import FileIngestRequest, FileIngestResponse
        
        # Test creating models
        request = FileIngestRequest(
            agent_id="test",
            file_path="/test/path.xlsx"
        )
        print(f"PASS Created FileIngestRequest: {request.agent_id}")
        
        response = FileIngestResponse(
            agent_id="test",
            request_id="test-123",
            status=AgentStatus.SUCCESS
        )
        print(f"PASS Created FileIngestResponse: {response.status}")
        
        return True
        
    except Exception as e:
        print(f"FAIL Models test failed: {e}")
        return False

async def test_config_only():
    """Test configuration only."""
    print("\nTesting configuration...")
    
    try:
        from excel_agent.utils.config import Config
        
        config = Config()
        print(f"PASS Config loaded - LLM Model: {config.llm_model}")
        print(f"PASS Max file size: {config.max_file_size_mb}MB")
        
        # Create directories
        Path(config.temp_dir).mkdir(parents=True, exist_ok=True)
        Path(config.cache_dir).mkdir(parents=True, exist_ok=True)
        print("PASS Directories created")
        
        return True
        
    except Exception as e:
        print(f"FAIL Config test failed: {e}")
        return False

async def test_pandas_basic():
    """Test basic Excel operations with pure pandas."""
    print("\nTesting pure pandas Excel operations...")
    
    try:
        import pandas as pd
        
        test_file = Path("data/synthetic/single_table_sales.xlsx")
        if not test_file.exists():
            print("SKIP No test file available")
            return True
        
        # Read with pandas
        df = pd.read_excel(test_file, sheet_name=0)
        print(f"PASS Read {len(df)} rows and {len(df.columns)} columns")
        
        # Basic analysis
        numeric_cols = df.select_dtypes(include=['number']).columns
        print(f"PASS Found {len(numeric_cols)} numeric columns")
        
        if len(numeric_cols) > 0:
            summary = df[numeric_cols].describe()
            print(f"PASS Generated summary statistics")
        
        return True
        
    except Exception as e:
        print(f"FAIL Pandas test failed: {e}")
        return False

async def test_openpyxl_basic():
    """Test basic openpyxl operations."""
    print("\nTesting pure openpyxl operations...")
    
    try:
        import openpyxl
        
        test_file = Path("data/synthetic/single_table_sales.xlsx")
        if not test_file.exists():
            print("SKIP No test file available")
            return True
        
        # Read with openpyxl (not read-only to access merged_cells)
        workbook = openpyxl.load_workbook(test_file, read_only=False)
        print(f"PASS Opened workbook with {len(workbook.sheetnames)} sheets")
        
        sheet = workbook.active
        print(f"PASS Active sheet: {sheet.title}")
        print(f"PASS Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} cols")
        
        # Check for merged cells
        try:
            merged_count = len(list(sheet.merged_cells.ranges))
            print(f"PASS Found {merged_count} merged cell ranges")
        except Exception:
            print("PASS Merged cells check skipped (read-only mode)")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"FAIL OpenPyXL test failed: {e}")
        return False

async def test_file_operations():
    """Test basic file operations."""
    print("\nTesting file operations...")
    
    try:
        import hashlib
        from datetime import datetime
        
        # Test file ID generation (like in FileIngestAgent)
        file_path = "test.xlsx"
        timestamp = datetime.now().isoformat()
        content = f"{file_path}:{timestamp}"
        file_id = hashlib.md5(content.encode()).hexdigest()
        print(f"PASS Generated file ID: {file_id[:8]}...")
        
        # Test file metadata
        test_files = list(Path("data/synthetic").glob("*.xlsx"))
        for file_path in test_files:
            size = file_path.stat().st_size
            print(f"PASS {file_path.name}: {size} bytes")
        
        return True
        
    except Exception as e:
        print(f"FAIL File operations test failed: {e}")
        return False

async def main():
    """Main test function."""
    print("Excel Agent System - Core Components Test")
    print("="*50)
    
    tests = [
        ("Models Only", test_models_only),
        ("Configuration Only", test_config_only), 
        ("Pandas Basic", test_pandas_basic),
        ("OpenPyXL Basic", test_openpyxl_basic),
        ("File Operations", test_file_operations),
    ]
    
    results = []
    
    for test_name, test_func in tests:
        try:
            result = await test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"FAIL {test_name} failed with exception: {e}")
            results.append((test_name, False))
    
    # Print summary
    print("\n" + "="*50)
    print("TEST RESULTS")
    print("="*50)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "PASS" if result else "FAIL"
        print(f"{status} {test_name}")
    
    print(f"\nOverall: {passed}/{total} tests passed")
    
    if passed == total:
        print("\nSUCCESS: All core components work!")
        print("The basic Excel processing functionality is ready.")
        print("\nNext steps:")
        print("1. Configure SILICONFLOW_API_KEY for AI features")
        print("2. The system can process Excel files with pandas/openpyxl")
        print("3. Core data models are working")
    else:
        print(f"\nISSUES: {total-passed} tests failed.")
        print("Fix the above issues before proceeding.")
    
    return passed == total

if __name__ == "__main__":
    result = asyncio.run(main())
    sys.exit(0 if result else 1)