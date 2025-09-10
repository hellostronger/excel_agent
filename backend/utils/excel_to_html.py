"""Excel to HTML conversion utilities."""

import pandas as pd
import openpyxl
from typing import Dict, List, Optional, Any
import html
from pathlib import Path


def excel_to_html(
    file_path: str, 
    sheet_name: Optional[str] = None,
    max_rows: int = 100,
    max_cols: int = 20,
    include_styling: bool = True
) -> Dict[str, Any]:
    """
    Convert Excel file to HTML format for preview with proper merged cell handling.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of sheet to convert (None for first sheet)
        max_rows: Maximum number of rows to include
        max_cols: Maximum number of columns to include
        include_styling: Whether to include basic styling
        
    Returns:
        Dictionary containing HTML content and metadata
    """
    try:
        # Try openpyxl approach first for better merged cell support
        try:
            return excel_to_html_openpyxl(file_path, sheet_name, max_rows, max_cols, include_styling)
        except Exception as openpyxl_error:
            print(f"OpenPyXL approach failed: {openpyxl_error}, falling back to pandas")
            # Fallback to pandas approach
            pass
        
        # Pandas fallback approach
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows, header=None)
        else:
            df = pd.read_excel(file_path, nrows=max_rows, header=None)
        
        # Limit columns if necessary
        if len(df.columns) > max_cols:
            df = df.iloc[:, :max_cols]
        
        # Get sheet names for metadata
        xl_file = pd.ExcelFile(file_path)
        sheet_names = xl_file.sheet_names
        
        # Handle missing values
        df = df.fillna('')
        
        # Generate HTML table
        html_content = generate_html_table_raw(df, include_styling=include_styling)
        
        # Prepare metadata
        metadata = {
            'total_rows': len(df),
            'total_cols': len(df.columns),
            'sheet_names': sheet_names,
            'current_sheet': sheet_name or sheet_names[0] if sheet_names else 'Sheet1',
            'columns': [f'Column_{i+1}' for i in range(len(df.columns))],
            'truncated': {
                'rows': len(df) >= max_rows,
                'cols': len(df.columns) >= max_cols
            }
        }
        
        return {
            'success': True,
            'html': html_content,
            'metadata': metadata,
            'preview_data': df.head(10).to_dict('records') if len(df) > 0 else []
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'html': f'<div class="error">Error loading Excel file: {html.escape(str(e))}</div>',
            'metadata': {}
        }


def excel_to_html_openpyxl(
    file_path: str,
    sheet_name: Optional[str] = None,
    max_rows: int = 100,
    max_cols: int = 20,
    include_styling: bool = True
) -> Dict[str, Any]:
    """Use openpyxl to handle merged cells properly."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(file_path, read_only=False, data_only=True)
    
    # Get worksheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        current_sheet = sheet_name
    else:
        ws = wb.active
        current_sheet = ws.title
    
    # Get the data range
    max_row = min(ws.max_row, max_rows) if ws.max_row else max_rows
    max_col = min(ws.max_column, max_cols) if ws.max_column else max_cols
    
    # Build merged cell map
    merged_ranges = {}
    for merged_range in ws.merged_ranges:
        min_row, min_col, max_row_range, max_col_range = merged_range.bounds
        for row in range(min_row, max_row_range + 1):
            for col in range(min_col, max_col_range + 1):
                if (row, col) not in merged_ranges:
                    merged_ranges[(row, col)] = {
                        'is_merged': True,
                        'is_top_left': (row == min_row and col == min_col),
                        'rowspan': max_row_range - min_row + 1,
                        'colspan': max_col_range - min_col + 1,
                        'value': ws.cell(min_row, min_col).value
                    }
    
    # Generate HTML table
    html_parts = []
    
    if include_styling:
        html_parts.append(get_excel_styling())
    
    html_parts.append('<table class="excel-preview">')
    
    for row in range(1, max_row + 1):
        html_parts.append('<tr>')
        
        for col in range(1, max_col + 1):
            cell_key = (row, col)
            
            # Skip cells that are part of merged range but not top-left
            if cell_key in merged_ranges and not merged_ranges[cell_key]['is_top_left']:
                continue
            
            cell = ws.cell(row, col)
            cell_value = cell.value or ''
            
            # Handle merged cells
            if cell_key in merged_ranges:
                merge_info = merged_ranges[cell_key]
                rowspan = f' rowspan="{merge_info["rowspan"]}"' if merge_info['rowspan'] > 1 else ''
                colspan = f' colspan="{merge_info["colspan"]}"' if merge_info['colspan'] > 1 else ''
                cell_value = merge_info['value'] or ''
            else:
                rowspan = colspan = ''
            
            # Format cell content
            if cell_value is None or cell_value == '':
                cell_content = '<span class="empty-cell">-</span>'
                cell_class = ''
            else:
                cell_content = html.escape(str(cell_value))
                # Check if numeric
                cell_class = ' class="numeric"' if isinstance(cell_value, (int, float)) else ''
                
                # Truncate long content
                if len(cell_content) > 100:
                    cell_content = cell_content[:97] + '...'
            
            # Add table cell
            tag = 'th' if row == 1 else 'td'
            html_parts.append(f'<{tag}{cell_class}{rowspan}{colspan} title="{html.escape(str(cell_value))}">{cell_content}</{tag}>')
        
        html_parts.append('</tr>')
    
    html_parts.append('</table>')
    
    # Prepare metadata
    metadata = {
        'total_rows': max_row,
        'total_cols': max_col,
        'sheet_names': wb.sheetnames,
        'current_sheet': current_sheet,
        'columns': [get_column_letter(i) for i in range(1, max_col + 1)],
        'truncated': {
            'rows': ws.max_row > max_rows if ws.max_row else False,
            'cols': ws.max_column > max_cols if ws.max_column else False
        },
        'merged_cells_count': len(ws.merged_ranges)
    }
    
    wb.close()
    
    return {
        'success': True,
        'html': ''.join(html_parts),
        'metadata': metadata,
        'preview_data': []  # Can be implemented later if needed
    }


def get_excel_styling() -> str:
    """Get CSS styling for Excel preview."""
    return """
<style>
.excel-preview {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    border-collapse: collapse;
    width: 100%;
    margin: 0;
    font-size: 12px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}

.excel-preview thead {
    background-color: #f8f9fa;
    position: sticky;
    top: 0;
    z-index: 10;
}

.excel-preview th {
    background-color: #e9ecef;
    border: 1px solid #dee2e6;
    padding: 8px 12px;
    text-align: center;
    font-weight: 600;
    color: #495057;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 200px;
}

.excel-preview td {
    border: 1px solid #dee2e6;
    padding: 6px 12px;
    vertical-align: middle;
    white-space: pre-wrap;
    word-wrap: break-word;
    max-width: 200px;
    overflow: hidden;
    text-overflow: ellipsis;
}

.excel-preview tbody tr:nth-child(even) {
    background-color: #f8f9fa;
}

.excel-preview tbody tr:hover {
    background-color: #e3f2fd;
}

.excel-preview .numeric {
    text-align: right;
    font-family: 'Courier New', monospace;
}

.excel-preview .empty-cell {
    color: #adb5bd;
    font-style: italic;
}

.empty-data, .error {
    padding: 20px;
    text-align: center;
    color: #6c757d;
    background-color: #f8f9fa;
    border: 1px solid #dee2e6;
    border-radius: 4px;
    margin: 10px 0;
}

.error {
    color: #dc3545;
    background-color: #f8d7da;
    border-color: #f5c6cb;
}
</style>
"""


def generate_html_table_raw(df: pd.DataFrame, include_styling: bool = True) -> str:
    """Generate HTML table from raw DataFrame data without column headers."""
    
    if df.empty:
        return '<div class="empty-data">No data available</div>'
    
    # Start HTML
    html_parts = []
    
    if include_styling:
        html_parts.append(get_excel_styling())
    
    # Start table
    html_parts.append('<table class="excel-preview">')
    
    # Table body (no separate header since we're treating first row as header)
    html_parts.append('<tbody>')
    
    for idx, row in df.iterrows():
        html_parts.append('<tr>')
        for col_idx, col in enumerate(df.columns):
            cell_value = row[col]
            cell_class = ""
            
            # Handle different data types
            if pd.isna(cell_value) or cell_value == '':
                cell_content = '<span class="empty-cell">-</span>'
            else:
                # Check if numeric
                if pd.api.types.is_numeric_dtype(type(cell_value)) and not pd.isna(cell_value):
                    cell_class = ' class="numeric"'
                    # Format numbers nicely
                    if isinstance(cell_value, float):
                        if cell_value.is_integer():
                            cell_content = str(int(cell_value))
                        else:
                            cell_content = f"{cell_value:.2f}" if abs(cell_value) < 1000000 else f"{cell_value:.2e}"
                    else:
                        cell_content = str(cell_value)
                else:
                    cell_content = html.escape(str(cell_value))
                
                # Truncate long content
                if len(cell_content) > 100:
                    cell_content = cell_content[:97] + '...'
            
            # Use th for first row, td for others
            tag = 'th' if idx == 0 else 'td'
            html_parts.append(f'<{tag}{cell_class} title="{html.escape(str(cell_value))}">{cell_content}</{tag}>')
        
        html_parts.append('</tr>')
    
    html_parts.append('</tbody></table>')
    
    return ''.join(html_parts)


def generate_html_table(df: pd.DataFrame, include_styling: bool = True) -> str:
    """Generate HTML table from pandas DataFrame with optional styling."""
    
    if df.empty:
        return '<div class="empty-data">No data available</div>'
    
    # Start HTML
    html_parts = []
    
    if include_styling:
        html_parts.append("""
<style>
.excel-preview {
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    border-collapse: collapse;
    width: 100%;
    margin: 0;
    font-size: 12px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
}

.excel-preview thead {
    background-color: #f8f9fa;
    position: sticky;
    top: 0;
    z-index: 10;
}

.excel-preview th {
    background-color: #e9ecef;
    border: 1px solid #dee2e6;
    padding: 8px 12px;
    text-align: left;
    font-weight: 600;
    color: #495057;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 200px;
}

.excel-preview td {
    border: 1px solid #dee2e6;
    padding: 6px 12px;
    vertical-align: top;
    white-space: pre-wrap;
    word-wrap: break-word;
    max-width: 200px;
    overflow: hidden;
    text-overflow: ellipsis;
}

.excel-preview tbody tr:nth-child(even) {
    background-color: #f8f9fa;
}

.excel-preview tbody tr:hover {
    background-color: #e3f2fd;
}

.excel-preview .numeric {
    text-align: right;
    font-family: 'Courier New', monospace;
}

.excel-preview .empty-cell {
    color: #adb5bd;
    font-style: italic;
}

.empty-data, .error {
    padding: 20px;
    text-align: center;
    color: #6c757d;
    background-color: #f8f9fa;
    border: 1px solid #dee2e6;
    border-radius: 4px;
    margin: 10px 0;
}

.error {
    color: #dc3545;
    background-color: #f8d7da;
    border-color: #f5c6cb;
}
</style>
""")
    
    # Start table
    html_parts.append('<table class="excel-preview">')
    
    # Table header
    html_parts.append('<thead><tr>')
    for col in df.columns:
        col_name = html.escape(str(col))
        if col_name == '' or col_name == 'nan':
            col_name = f'Column_{df.columns.get_loc(col) + 1}'
        html_parts.append(f'<th title="{col_name}">{col_name}</th>')
    html_parts.append('</tr></thead>')
    
    # Table body
    html_parts.append('<tbody>')
    
    for idx, row in df.iterrows():
        html_parts.append('<tr>')
        for col in df.columns:
            cell_value = row[col]
            cell_class = ""
            
            # Handle different data types
            if pd.isna(cell_value) or cell_value == '':
                cell_content = '<span class="empty-cell">-</span>'
            else:
                # Check if numeric
                if pd.api.types.is_numeric_dtype(df[col]) and not pd.isna(cell_value):
                    cell_class = ' class="numeric"'
                    # Format numbers nicely
                    if isinstance(cell_value, float):
                        if cell_value.is_integer():
                            cell_content = str(int(cell_value))
                        else:
                            cell_content = f"{cell_value:.2f}" if abs(cell_value) < 1000000 else f"{cell_value:.2e}"
                    else:
                        cell_content = str(cell_value)
                else:
                    cell_content = html.escape(str(cell_value))
                
                # Truncate long content
                if len(cell_content) > 100:
                    cell_content = cell_content[:97] + '...'
            
            html_parts.append(f'<td{cell_class} title="{html.escape(str(cell_value))}">{cell_content}</td>')
        
        html_parts.append('</tr>')
    
    html_parts.append('</tbody></table>')
    
    return ''.join(html_parts)


def get_excel_sheets(file_path: str) -> List[str]:
    """Get list of sheet names from Excel file."""
    try:
        xl_file = pd.ExcelFile(file_path)
        return xl_file.sheet_names
    except Exception:
        return []


def excel_sheet_to_html(
    file_path: str,
    sheet_name: str,
    start_row: int = 0,
    end_row: Optional[int] = None,
    start_col: int = 0,
    end_col: Optional[int] = None
) -> Dict[str, Any]:
    """
    Convert specific range of Excel sheet to HTML.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet
        start_row: Starting row index (0-based)
        end_row: Ending row index (None for all rows)
        start_col: Starting column index (0-based)  
        end_col: Ending column index (None for all columns)
    
    Returns:
        Dictionary with HTML content and metadata
    """
    try:
        # Read specific sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # Apply range selection
        if end_row is not None:
            df = df.iloc[start_row:end_row]
        else:
            df = df.iloc[start_row:]
            
        if end_col is not None:
            df = df.iloc[:, start_col:end_col]
        else:
            df = df.iloc[:, start_col:]
        
        # Generate HTML
        html_content = generate_html_table(df, include_styling=True)
        
        return {
            'success': True,
            'html': html_content,
            'metadata': {
                'sheet_name': sheet_name,
                'rows': len(df),
                'columns': len(df.columns),
                'range': f'{start_row}:{end_row or "end"}x{start_col}:{end_col or "end"}'
            }
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'html': f'<div class="error">Error loading sheet {sheet_name}: {html.escape(str(e))}</div>'
        }


def get_excel_info(file_path: str) -> Dict[str, Any]:
    """Get comprehensive information about Excel file."""
    try:
        xl_file = pd.ExcelFile(file_path)
        file_info = {
            'file_name': Path(file_path).name,
            'file_size': Path(file_path).stat().st_size,
            'sheet_count': len(xl_file.sheet_names),
            'sheets': []
        }
        
        # Get info for each sheet
        for sheet_name in xl_file.sheet_names:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1)
                sheet_info = {
                    'name': sheet_name,
                    'columns': len(df.columns),
                    'column_names': list(df.columns)
                }
                
                # Get row count (more efficient way)
                df_full = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_info['rows'] = len(df_full)
                
                file_info['sheets'].append(sheet_info)
            except Exception as e:
                file_info['sheets'].append({
                    'name': sheet_name,
                    'error': str(e)
                })
        
        return {
            'success': True,
            'info': file_info
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }