"""File Ingest Agent for loading and parsing Excel files."""

import os
import hashlib
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from .base import BaseAgent
from ..models.agents import FileIngestRequest, FileIngestResponse
from ..models.base import FileMetadata, AgentRequest, AgentResponse, AgentStatus
from ..utils.config import config


class FileIngestAgent(BaseAgent):
    """Enhanced agent for loading and parsing Excel files with ST-Raptor optimizations."""
    
    def __init__(self):
        super().__init__(
            name="FileIngestAgent",
            description="Loads and parses Excel files, extracts sheet names and metadata, creates feature trees",
            mcp_capabilities=["excel_tools", "file_management"]
        )
        self.file_store = {}  # In-memory store for file metadata
        self.config = get_config()
        self.cache_manager = get_cache_manager()
        self.embedding_agent = EmbeddingAgent()
    
    async def process(self, request: AgentRequest) -> AgentResponse:
        """Process file ingestion request."""
        if not isinstance(request, FileIngestRequest):
            return self.create_error_response(
                request, 
                f"Invalid request type. Expected FileIngestRequest, got {type(request)}"
            )
        
        try:
            # Validate file path
            file_path = Path(request.file_path)
            if not file_path.exists():
                return self.create_error_response(
                    request,
                    f"File does not exist: {request.file_path}"
                )
            
            if not file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
                return self.create_error_response(
                    request,
                    f"Unsupported file format: {file_path.suffix}"
                )
            
            # Check file size
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            if file_size_mb > self.config.max_file_size_mb:
                return self.create_error_response(
                    request,
                    f"File too large: {file_size_mb:.1f}MB > {self.config.max_file_size_mb}MB"
                )
            
            # Generate file ID for caching
            file_id = self._generate_file_id(file_path)
            
            # Check file size
            file_size = file_path.stat().st_size
            if file_size > self.config.max_file_size_mb * 1024 * 1024:
                return self.create_error_response(
                    request,
                    f"File too large: {file_size / (1024*1024):.1f}MB > {self.config.max_file_size_mb}MB"
                )
            
            # Generate file ID
            file_id = self._generate_file_id(request.file_path)
            
            # Parse Excel file
            metadata = await self._parse_excel_file(file_path, file_id)
            
            # Use MCP tools for enhanced file processing if available
            mcp_result = await self._enhance_with_mcp(file_path, metadata)
            if mcp_result:
                metadata = mcp_result
            
            # Store metadata
            self.file_store[file_id] = metadata
            
            self.logger.info(f"Successfully ingested file: {file_path.name}")
            
            # Create response
            return FileIngestResponse(
                agent_id=self.name,
                request_id=request.request_id,
                status=AgentStatus.SUCCESS,
                file_id=file_id,
                sheets=metadata.sheets,
                result={
                    "file_id": file_id,
                    "file_path": str(file_path),
                    "sheets": metadata.sheets,
                    "metadata": metadata.model_dump()
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error processing file ingestion: {e}")
            return self.create_error_response(request, str(e))
    
    async def _process_file_original(self, request: AgentRequest, file_path: Path, file_id: str) -> AgentResponse:
        """Original file processing pipeline (preserved for backward compatibility)."""
        try:
            # Parse Excel file
            metadata = await self._parse_excel_file(file_path, file_id)
            
            # Use MCP tools for enhanced file processing if available
            mcp_result = await self._enhance_with_mcp(file_path, metadata)
            if mcp_result:
                metadata = mcp_result
            
            # Store metadata
            self.file_store[file_id] = metadata
            
            self.logger.info(f"Successfully ingested file: {file_path.name}")
            
            # Create response using original format
            from ..models.agents import FileIngestResponse
            from ..models.base import AgentStatus
            
            return FileIngestResponse(
                agent_id=self.name,
                request_id=request.request_id,
                status=AgentStatus.SUCCESS,
                file_id=file_id,
                sheets=metadata.sheets,
                result={
                    "file_id": file_id,
                    "file_path": str(file_path),
                    "sheets": metadata.sheets,
                    "metadata": metadata.model_dump() if hasattr(metadata, 'model_dump') else metadata.__dict__
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error in original file processing: {e}")
            return self.create_error_response(request, str(e))
    
    async def _process_file_with_st_raptor(self, request: AgentRequest, file_path: Path, file_id: str) -> AgentResponse:
        """Enhanced file processing with ST-Raptor optimizations (feature trees, embeddings, caching)."""
        try:
            # Parse Excel file metadata
            metadata = await self._parse_excel_file(file_path, file_id)
            
            # Create feature tree
            feature_tree = await self._create_feature_tree(file_path, file_id, metadata)
            
            # Generate embeddings if enabled
            embedding_dict = None
            if self.config.enable_embedding_cache:
                embedding_dict = self.embedding_agent.create_embedding_dict(feature_tree)
                self.cache_manager.save_embedding_cache(embedding_dict, file_id)
            
            # Store enhanced data (ST-Raptor format)
            self.file_store[file_id] = {
                "metadata": metadata,
                "feature_tree": feature_tree,
                "embedding_dict": embedding_dict,
                "processing_mode": "st_raptor"
            }
            
            # Also check cache for ST-Raptor processing
            if self.config.enable_cache:
                cached_result = self._load_from_cache(file_id)
                if cached_result:
                    self.logger.info(f"Loaded ST-Raptor data from cache: {file_id}")
                    return cached_result
            
            # Save to cache
            if self.config.enable_cache:
                self._save_to_cache(file_id, None)  # Save the enhanced data
            
            self.logger.info(f"Successfully processed file with feature tree: {file_path.name}")
            
            # Create enhanced response
            return AgentResponse(
                request_id=request.request_id,
                agent_name=self.name,
                status="completed",
                data={
                    "file_id": file_id,
                    "file_path": str(file_path),
                    "metadata": metadata.model_dump() if hasattr(metadata, 'model_dump') else metadata.__dict__,
                    "feature_tree_stats": feature_tree.get_statistics(),
                    "embedding_stats": {
                        "count": embedding_dict.get("count", 0) if embedding_dict else 0,
                        "dimension": embedding_dict.get("dimension", 0) if embedding_dict else 0
                    },
                    "cached": True
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error processing file with tree: {e}")
            return AgentResponse(
                request_id=request.request_id,
                agent_name=self.name,
                status="error",
                error_message=str(e)
            )
    
    async def _create_feature_tree(self, file_path: Path, file_id: str, metadata: FileMetadata) -> FeatureTree:
        """Create feature tree from Excel file (inspired by ST-Raptor)."""
        try:
            tree = FeatureTree()
            tree.table_id = file_id
            tree.file_path = str(file_path)
            tree.set_metadata(metadata.model_dump() if hasattr(metadata, 'model_dump') else metadata.__dict__)
            
            # Create root node
            root = IndexNode(value=f"Excel_File_{file_path.stem}")
            
            # Process each sheet
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in metadata.sheets:
                sheet = workbook[sheet_name]
                sheet_node = IndexNode(value=f"Sheet_{sheet_name}")
                
                # Analyze sheet structure
                sheet_stats = self._analyze_sheet_structure(sheet)
                
                # Add column nodes
                if sheet.max_column and sheet.max_column > 0:
                    for col_idx in range(1, min(sheet.max_column + 1, 50)):  # Limit to first 50 columns
                        col_letter = get_column_letter(col_idx)
                        col_node = IndexNode(value=f"Column_{col_letter}")
                        
                        # Add cell values as body items
                        cell_values = []
                        for row_idx in range(1, min(sheet.max_row + 1, 100)):  # Limit to first 100 rows
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                cell_values.append(TreeNode(value=str(cell.value)))
                        
                        for cell_node in cell_values[:20]:  # Limit to first 20 values per column
                            col_node.add_body_node(cell_node)
                        
                        sheet_node.add_child(col_node)
                
                # Add merged cell information if exists
                if sheet.merged_cells:
                    merged_node = IndexNode(value="Merged_Cells")
                    for merged_range in list(sheet.merged_cells)[:10]:  # Limit to first 10 merged ranges
                        range_node = TreeNode(value=str(merged_range))
                        merged_node.add_body_node(range_node)
                    sheet_node.add_child(merged_node)
                
                root.add_child(sheet_node)
            
            workbook.close()
            tree.set_root(root)
            
            # Add schema information
            schema_info = {
                "file_type": "excel",
                "sheets": metadata.sheets,
                "structure_complexity": self._assess_structure_complexity(metadata),
                "has_merged_cells": metadata.has_merged_cells,
                "has_formulas": metadata.has_formulas
            }
            tree.set_schema_info(schema_info)
            
            return tree
            
        except Exception as e:
            self.logger.error(f"Failed to create feature tree: {e}")
            # Create minimal tree as fallback
            tree = FeatureTree()
            tree.table_id = file_id
            tree.file_path = str(file_path)
            root = IndexNode(value="Error_Processing_File")
            tree.set_root(root)
            return tree
    
    def _analyze_sheet_structure(self, sheet) -> Dict[str, Any]:
        """Analyze sheet structure for feature tree creation."""
        stats = {
            "row_count": sheet.max_row or 0,
            "column_count": sheet.max_column or 0,
            "merged_cell_count": len(sheet.merged_cells) if sheet.merged_cells else 0,
            "has_data": False,
            "data_density": 0.0
        }
        
        # Sample cells to estimate data density
        if sheet.max_row and sheet.max_column:
            sample_size = min(100, sheet.max_row * sheet.max_column)
            filled_cells = 0
            
            for row in sheet.iter_rows(max_row=min(10, sheet.max_row), 
                                     max_col=min(10, sheet.max_column)):
                for cell in row:
                    if cell.value is not None:
                        filled_cells += 1
                        stats["has_data"] = True
            
            if sample_size > 0:
                stats["data_density"] = filled_cells / min(100, sample_size)
        
        return stats
    
    def _assess_structure_complexity(self, metadata: FileMetadata) -> str:
        """Assess the structural complexity of the Excel file."""
        complexity_score = 0
        
        # Multiple sheets add complexity
        if len(metadata.sheets) > 1:
            complexity_score += 1
        
        # Large files add complexity
        if metadata.total_rows > 1000 or metadata.total_columns > 50:
            complexity_score += 1
        
        # Special features add complexity
        if metadata.has_merged_cells:
            complexity_score += 1
        if metadata.has_formulas:
            complexity_score += 1
        if metadata.has_charts:
            complexity_score += 1
        
        if complexity_score >= 4:
            return "high"
        elif complexity_score >= 2:
            return "medium"
        else:
            return "low"
    
    def _generate_file_id(self, file_path: Path) -> str:
        """Generate file ID based on path and modification time."""
        file_stat = file_path.stat()
        content = f"{file_path.absolute()}:{file_stat.st_mtime}:{file_stat.st_size}"
        return hashlib.md5(content.encode()).hexdigest()
    
    def _load_from_cache(self, file_id: str) -> Optional[AgentResponse]:
        """Load processed file data from cache."""
        try:
            # Try loading metadata from cache
            cached_metadata = self.cache_manager.load_metadata_cache(file_id)
            if not cached_metadata:
                return None
            
            # Try loading feature tree from cache
            feature_tree = self.cache_manager.load_tree_cache(file_id)
            if not feature_tree:
                return None
            
            # Try loading embeddings from cache
            embedding_dict = self.cache_manager.load_embedding_cache(file_id) if self.config.enable_embedding_cache else None
            
            # Store in memory
            self.file_store[file_id] = {
                "metadata": cached_metadata,
                "feature_tree": feature_tree,
                "embedding_dict": embedding_dict
            }
            
            return AgentResponse(
                request_id=file_id,  # Use file_id as placeholder
                agent_name=self.name,
                status="completed",
                data={
                    "file_id": file_id,
                    "file_path": cached_metadata.get("file_path", "unknown"),
                    "metadata": cached_metadata,
                    "feature_tree_stats": feature_tree.get_statistics(),
                    "embedding_stats": {
                        "count": embedding_dict.get("count", 0) if embedding_dict else 0,
                        "dimension": embedding_dict.get("dimension", 0) if embedding_dict else 0
                    },
                    "cached": True
                }
            )
            
        except Exception as e:
            self.logger.warning(f"Failed to load from cache: {e}")
            return None
    
    def _save_to_cache(self, file_id: str, response: AgentResponse):
        """Save processed file data to cache."""
        try:
            if file_id not in self.file_store:
                return
            
            stored_data = self.file_store[file_id]
            
            # Save metadata
            if "metadata" in stored_data:
                self.cache_manager.save_metadata_cache(stored_data["metadata"], file_id)
            
            # Save feature tree
            if "feature_tree" in stored_data:
                self.cache_manager.save_tree_cache(stored_data["feature_tree"], file_id)
            
            # Embeddings are already saved during processing
            
        except Exception as e:
            self.logger.warning(f"Failed to save to cache: {e}")
    
    async def _parse_excel_file(self, file_path: Path, file_id: str) -> FileMetadata:
        """Parse Excel file and extract metadata."""
        try:
            # Load with openpyxl for rich metadata
            workbook = openpyxl.load_workbook(
                file_path, 
                read_only=False,
                keep_vba=True,
                data_only=False,
                keep_links=True
            )
            
            sheets = []
            total_rows = 0
            total_columns = 0
            has_merged_cells = False
            has_formulas = False
            has_charts = False
            has_images = False
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheets.append(sheet_name)
                
                # Count rows and columns
                if sheet.max_row:
                    total_rows += sheet.max_row
                if sheet.max_column:
                    total_columns += sheet.max_column
                
                # Check for merged cells
                if sheet.merged_cells:
                    has_merged_cells = True
                
                # Check for formulas
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            has_formulas = True
                            break
                    if has_formulas:
                        break
                
                # Check for charts
                if hasattr(sheet, '_charts') and sheet._charts:
                    has_charts = True
                
                # Check for images
                if hasattr(sheet, '_images') and sheet._images:
                    has_images = True
            
            workbook.close()
            
            # Create metadata
            metadata = FileMetadata(
                file_id=file_id,
                original_filename=file_path.name,
                file_path=str(file_path.absolute()),
                file_size=file_path.stat().st_size,
                upload_timestamp=datetime.now(),
                sheets=sheets,
                total_rows=total_rows,
                total_columns=total_columns,
                has_merged_cells=has_merged_cells,
                has_formulas=has_formulas,
                has_charts=has_charts,
                has_images=has_images
            )
            
            return metadata
            
        except Exception as e:
            # Fallback to pandas for basic info
            self.logger.warning(f"openpyxl failed, falling back to pandas: {e}")
            
            try:
                # Read all sheets with pandas
                excel_file = pd.ExcelFile(file_path)
                sheets = excel_file.sheet_names
                
                total_rows = 0
                total_columns = 0
                
                for sheet_name in sheets:
                    df = excel_file.parse(sheet_name)
                    total_rows += len(df)
                    total_columns += len(df.columns)
                
                metadata = FileMetadata(
                    file_id=file_id,
                    original_filename=file_path.name,
                    file_path=str(file_path.absolute()),
                    file_size=file_path.stat().st_size,
                    upload_timestamp=datetime.now(),
                    sheets=sheets,
                    total_rows=total_rows,
                    total_columns=total_columns,
                    has_merged_cells=False,  # Unknown with pandas
                    has_formulas=False,     # Unknown with pandas
                    has_charts=False,       # Unknown with pandas
                    has_images=False        # Unknown with pandas
                )
                
                return metadata
                
            except Exception as e2:
                raise Exception(f"Failed to parse Excel file with both openpyxl and pandas: {e2}")
    
    def _generate_file_id(self, file_path: str) -> str:
        """Generate a unique file ID based on file path and timestamp."""
        timestamp = datetime.now().isoformat()
        content = f"{file_path}:{timestamp}"
        return hashlib.md5(content.encode()).hexdigest()
    
    def get_file_metadata(self, file_id: str) -> Optional[FileMetadata]:
        """Get file metadata by file ID."""
        return self.file_store.get(file_id)
    
    def list_files(self) -> List[FileMetadata]:
        """List all ingested files."""
        return list(self.file_store.values())
    
    def remove_file(self, file_id: str) -> bool:
        """Remove file from store."""
        if file_id in self.file_store:
            del self.file_store[file_id]
            return True
        return False
    
    async def _enhance_with_mcp(self, file_path: Path, metadata: FileMetadata) -> Optional[FileMetadata]:
        """Enhance file metadata using MCP tools."""
        try:
            # Use MCP Excel tools for enhanced analysis
            excel_info = await self.call_mcp_tool("read_excel_file", {
                "file_path": str(file_path)
            })
            
            if excel_info and not excel_info.get("error"):
                # Create backup using MCP file management
                backup_result = await self.call_mcp_tool("create_backup", {
                    "file_path": str(file_path)
                })
                
                # Enhance metadata with MCP results
                enhanced_metadata = metadata.model_copy()
                
                if excel_info.get("sheets"):
                    enhanced_metadata.sheets = list(excel_info["sheets"].keys())
                    
                    # Add sheet-specific information
                    total_rows = 0
                    total_columns = 0
                    for sheet_info in excel_info["sheets"].values():
                        if "shape" in sheet_info:
                            total_rows += sheet_info["shape"][0]
                            total_columns += sheet_info["shape"][1]
                    
                    enhanced_metadata.total_rows = total_rows
                    enhanced_metadata.total_columns = total_columns
                
                # Log backup status
                if backup_result and not backup_result.get("error"):
                    self.logger.info(f"Created backup: {backup_result.get('backup_file')}")
                
                return enhanced_metadata
            
            return None
            
        except Exception as e:
            self.logger.warning(f"MCP enhancement failed: {e}")
            return None