"""
Tree Structure Analysis and Intelligent Partitioning

基于ST-Raptor的树状结构分析和智能分片策略，实现Excel表格的层次化理解和处理。
"""

import logging
from typing import Dict, Any, List, Optional, Tuple, Union
from pathlib import Path
from dataclasses import dataclass
from enum import Enum
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from .utils import setup_logging
from .embedding_engine import get_embedding_engine

# Setup logging
logger = setup_logging()


class TableType(Enum):
    """表格类型枚举，参考ST-Raptor分类"""
    LIST = "list"           # T_LIST: 列表型表格
    ATTRIBUTE = "attribute" # T_ATTR: 属性型表格  
    SEMI = "semi"          # T_SEMI: 半结构化表格
    MIXED = "mixed"        # T_MIX: 混合型表格
    SIMPLE = "simple"      # 简单表格
    UNKNOWN = "unknown"    # 未知类型


class SchemaDirection(Enum):
    """Schema方向枚举"""
    TOP = "top"      # SCHEMA_TOP: 表头在顶部
    LEFT = "left"    # SCHEMA_LEFT: 表头在左侧
    UNKNOWN = "unknown"


@dataclass
class TableRegion:
    """表格区域定义"""
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    region_type: str  # header, data, summary, calculation
    confidence: float
    content_sample: List[str] = None


@dataclass
class TreeNode:
    """树结构节点，参考ST-Raptor的TreeNode"""
    value: Any
    node_type: str  # index, body, data
    children: List['TreeNode'] = None
    parent: 'TreeNode' = None
    position: Tuple[int, int, int, int] = None  # (start_row, start_col, end_row, end_col)
    metadata: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.children is None:
            self.children = []
        if self.metadata is None:
            self.metadata = {}


class IntelligentTablePartitioner:
    """
    智能表格分片器
    
    参考ST-Raptor的分片算法，实现基于结构理解的表格自动分片。
    """
    
    def __init__(self):
        self.small_table_threshold = {"rows": 8, "cols": 8}
        self.large_table_threshold = {"rows": 50, "cols": 20}
        self.embedding_engine = None
        
    def _get_embedding_engine(self):
        """延迟加载embedding引擎"""
        if self.embedding_engine is None:
            try:
                self.embedding_engine = get_embedding_engine()
            except Exception as e:
                logger.warning(f"Embedding engine not available: {e}")
                self.embedding_engine = None
        return self.embedding_engine
    
    async def analyze_and_partition(self, file_path: str) -> Dict[str, Any]:
        """
        分析并分片Excel文件
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            分片分析结果
        """
        try:
            path = Path(file_path)
            
            if path.suffix.lower() == '.xlsx':
                return await self._process_xlsx_file(file_path)
            elif path.suffix.lower() == '.xls':
                return await self._process_xls_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {path.suffix}")
                
        except Exception as e:
            logger.error(f"Table partitioning failed: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "partitions": []
            }
    
    async def _process_xlsx_file(self, file_path: str) -> Dict[str, Any]:
        """处理.xlsx文件"""
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        
        try:
            result = {
                "success": True,
                "file_path": file_path,
                "sheets": [],
                "processing_strategy": {},
                "partitions": []
            }
            
            for worksheet in workbook.worksheets:
                sheet_result = await self._process_worksheet(worksheet)
                result["sheets"].append(sheet_result)
                
                # 收集分片
                if sheet_result.get("partitions"):
                    for partition in sheet_result["partitions"]:
                        partition["source_sheet"] = worksheet.title
                        result["partitions"].append(partition)
            
            # 确定整体处理策略
            result["processing_strategy"] = self._determine_processing_strategy(result["sheets"])
            
            return result
            
        finally:
            workbook.close()
    
    async def _process_worksheet(self, worksheet) -> Dict[str, Any]:
        """处理单个工作表"""
        
        sheet_info = {
            "sheet_name": worksheet.title,
            "dimensions": {
                "rows": worksheet.max_row,
                "cols": worksheet.max_column
            },
            "table_type": TableType.UNKNOWN.value,
            "schema_direction": SchemaDirection.UNKNOWN.value,
            "regions": [],
            "partitions": [],
            "complexity_level": "medium",
            "processing_recommendation": {}
        }
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 步骤1: 判断是否为最小结构单元
        min_structure = self._match_minimal_structure(worksheet)
        if min_structure:
            sheet_info["table_type"] = TableType.SIMPLE.value
            sheet_info["partitions"] = [min_structure]
            sheet_info["complexity_level"] = "low"
            return sheet_info
        
        # 步骤2: 判断是否为小表格
        if self._is_small_table(max_row, max_col):
            sheet_info["complexity_level"] = "low"
            sheet_info["processing_recommendation"]["strategy"] = "direct_processing"
            sheet_info["processing_recommendation"]["use_vlm"] = True
            
            # 直接分析小表格
            direct_analysis = await self._analyze_small_table(worksheet)
            sheet_info.update(direct_analysis)
            return sheet_info
        
        # 步骤3: 检测Schema方向
        schema_direction = await self._detect_schema_direction(worksheet)
        sheet_info["schema_direction"] = schema_direction.value
        
        # 步骤4: 检测表格类型
        table_type = await self._classify_table_type(worksheet, schema_direction)
        sheet_info["table_type"] = table_type.value
        
        # 步骤5: 区域分析
        regions = await self._detect_table_regions(worksheet, schema_direction)
        sheet_info["regions"] = [self._region_to_dict(region) for region in regions]
        
        # 步骤6: 智能分片
        partitions = await self._intelligent_partitioning(worksheet, schema_direction, table_type, regions)
        sheet_info["partitions"] = partitions
        
        # 步骤7: 复杂度评估和处理建议
        complexity = self._assess_complexity(worksheet, partitions, regions)
        sheet_info["complexity_level"] = complexity["level"]
        sheet_info["processing_recommendation"] = complexity["recommendation"]
        
        return sheet_info
    
    def _match_minimal_structure(self, worksheet) -> Optional[Dict[str, Any]]:
        """
        匹配最小结构单元
        
        参考ST-Raptor的match_min_table_structure函数
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 检查是否为单个单元格或很小的结构
        if max_row <= 2 and max_col <= 2:
            # 获取第一个单元格
            first_cell = worksheet.cell(row=1, column=1)
            
            if max_row == 1 and max_col == 1:
                # 单个单元格
                return {
                    "type": "minimal_single",
                    "content": {str(first_cell.value): None},
                    "position": (1, 1, 1, 1),
                    "confidence": 1.0
                }
            
            elif max_row == 1 and max_col == 2:
                # 一行两列
                second_cell = worksheet.cell(row=1, column=2)
                return {
                    "type": "minimal_pair",
                    "content": {str(first_cell.value): str(second_cell.value)},
                    "position": (1, 1, 1, 2),
                    "confidence": 0.9
                }
                
            elif max_row == 2 and max_col == 1:
                # 两行一列
                second_cell = worksheet.cell(row=2, column=1)
                return {
                    "type": "minimal_pair",
                    "content": {str(first_cell.value): str(second_cell.value)},
                    "position": (1, 1, 2, 1),
                    "confidence": 0.9
                }
        
        return None
    
    def _is_small_table(self, rows: int, cols: int) -> bool:
        """判断是否为小表格"""
        return (rows <= self.small_table_threshold["rows"] and 
                cols <= self.small_table_threshold["cols"])
    
    async def _analyze_small_table(self, worksheet) -> Dict[str, Any]:
        """分析小表格"""
        
        analysis = {
            "processing_recommendation": {
                "strategy": "direct_vlm_processing",
                "reasoning": "Small table suitable for direct VLM analysis"
            }
        }
        
        # 提取所有内容
        content = []
        for row in worksheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            content.append(row_data)
        
        analysis["raw_content"] = content
        
        # 简单的结构分析
        if content:
            first_row = content[0]
            # 检查第一行是否像表头
            non_empty_first_row = [cell for cell in first_row if cell.strip()]
            if len(non_empty_first_row) > 1:
                analysis["likely_has_header"] = True
                analysis["table_type"] = TableType.LIST.value
            else:
                analysis["likely_has_header"] = False
                analysis["table_type"] = TableType.ATTRIBUTE.value
        
        return analysis
    
    async def _detect_schema_direction(self, worksheet) -> SchemaDirection:
        """
        检测Schema方向
        
        参考ST-Raptor的Schema方向检测逻辑
        """
        max_row = min(worksheet.max_row, 10)  # 只检测前10行
        max_col = min(worksheet.max_column, 10)  # 只检测前10列
        
        # 检测顶部Schema的特征
        top_score = 0.0
        first_row_content = []
        
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col)
            if cell.value:
                first_row_content.append(str(cell.value).strip().lower())
        
        if first_row_content:
            # 检查表头关键词
            header_keywords = ["id", "name", "date", "time", "type", "status", "amount", "value", 
                             "编号", "名称", "日期", "时间", "类型", "状态", "金额", "数值", "总计"]
            
            header_matches = sum(1 for content in first_row_content 
                               if any(keyword in content for keyword in header_keywords))
            
            if first_row_content:
                top_score = header_matches / len(first_row_content)
                
                # 检查文本比例（表头通常是文本）
                text_ratio = sum(1 for content in first_row_content 
                               if not self._is_numeric(content)) / len(first_row_content)
                top_score = (top_score + text_ratio) / 2
        
        # 检测左侧Schema的特征
        left_score = 0.0
        first_col_content = []
        
        for row in range(1, max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            if cell.value:
                first_col_content.append(str(cell.value).strip().lower())
        
        if first_col_content:
            header_matches = sum(1 for content in first_col_content 
                               if any(keyword in content for keyword in header_keywords))
            
            if first_col_content:
                left_score = header_matches / len(first_col_content)
                
                # 检查文本比例
                text_ratio = sum(1 for content in first_col_content 
                               if not self._is_numeric(content)) / len(first_col_content)
                left_score = (left_score + text_ratio) / 2
        
        # 决定方向
        if top_score > left_score and top_score > 0.3:
            return SchemaDirection.TOP
        elif left_score > top_score and left_score > 0.3:
            return SchemaDirection.LEFT
        elif top_score > 0.2 or left_score > 0.2:
            return SchemaDirection.TOP  # 默认为顶部
        else:
            return SchemaDirection.UNKNOWN
    
    def _is_numeric(self, value: str) -> bool:
        """判断字符串是否为数值"""
        try:
            float(value.replace(',', '').replace('%', '').replace('$', ''))
            return True
        except (ValueError, AttributeError):
            return False
    
    async def _classify_table_type(self, worksheet, schema_direction: SchemaDirection) -> TableType:
        """
        分类表格类型
        
        参考ST-Raptor的表格分类逻辑
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 基本规则
        if max_row <= 3 and max_col <= 3:
            return TableType.SIMPLE
        
        # 检测合并单元格模式
        merge_count = len(worksheet.merged_cells.ranges)
        merge_ratio = merge_count / (max_row * max_col) if max_row * max_col > 0 else 0
        
        if schema_direction == SchemaDirection.TOP:
            # 顶部Schema通常是列表型表格
            if merge_ratio < 0.1:
                return TableType.LIST
            else:
                return TableType.SEMI
        
        elif schema_direction == SchemaDirection.LEFT:
            # 左侧Schema通常是属性型表格
            if merge_ratio < 0.1:
                return TableType.ATTRIBUTE
            else:
                return TableType.SEMI
        
        else:
            # 未知Schema方向
            if merge_ratio > 0.2:
                return TableType.MIXED
            else:
                # 通过宽高比判断
                if max_col > max_row * 1.5:
                    return TableType.LIST
                elif max_row > max_col * 1.5:
                    return TableType.ATTRIBUTE
                else:
                    return TableType.MIXED
    
    async def _detect_table_regions(self, worksheet, schema_direction: SchemaDirection) -> List[TableRegion]:
        """检测表格区域"""
        
        regions = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if schema_direction == SchemaDirection.TOP:
            # 顶部Schema - 水平分割
            
            # 检测表头区域高度
            header_height = 1
            for row in range(1, min(5, max_row) + 1):
                row_content = []
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        row_content.append(str(cell.value).strip())
                
                if row_content:
                    # 检查是否像表头
                    text_ratio = sum(1 for content in row_content if not self._is_numeric(content)) / len(row_content)
                    if text_ratio > 0.7:  # 大部分是文本
                        header_height = row
                    else:
                        break
            
            # 表头区域
            if header_height >= 1:
                regions.append(TableRegion(
                    start_row=1, start_col=1,
                    end_row=header_height, end_col=max_col,
                    region_type="header",
                    confidence=0.8,
                    content_sample=self._extract_region_sample(worksheet, 1, 1, header_height, max_col)
                ))
            
            # 数据区域
            if header_height < max_row:
                regions.append(TableRegion(
                    start_row=header_height + 1, start_col=1,
                    end_row=max_row, end_col=max_col,
                    region_type="data",
                    confidence=0.9,
                    content_sample=self._extract_region_sample(worksheet, header_height + 1, 1, max_row, max_col)
                ))
        
        elif schema_direction == SchemaDirection.LEFT:
            # 左侧Schema - 垂直分割
            
            # 检测表头区域宽度
            header_width = 1
            for col in range(1, min(5, max_col) + 1):
                col_content = []
                for row in range(1, max_row + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        col_content.append(str(cell.value).strip())
                
                if col_content:
                    text_ratio = sum(1 for content in col_content if not self._is_numeric(content)) / len(col_content)
                    if text_ratio > 0.7:
                        header_width = col
                    else:
                        break
            
            # 表头区域
            if header_width >= 1:
                regions.append(TableRegion(
                    start_row=1, start_col=1,
                    end_row=max_row, end_col=header_width,
                    region_type="header",
                    confidence=0.8,
                    content_sample=self._extract_region_sample(worksheet, 1, 1, max_row, header_width)
                ))
            
            # 数据区域
            if header_width < max_col:
                regions.append(TableRegion(
                    start_row=1, start_col=header_width + 1,
                    end_row=max_row, end_col=max_col,
                    region_type="data",
                    confidence=0.9,
                    content_sample=self._extract_region_sample(worksheet, 1, header_width + 1, max_row, max_col)
                ))
        
        return regions
    
    def _extract_region_sample(self, worksheet, start_row: int, start_col: int, 
                              end_row: int, end_col: int, max_samples: int = 5) -> List[str]:
        """提取区域内容样本"""
        
        samples = []
        sample_count = 0
        
        for row in range(start_row, min(end_row + 1, start_row + 3)):  # 最多3行
            for col in range(start_col, min(end_col + 1, start_col + 5)):  # 最多5列
                if sample_count >= max_samples:
                    break
                    
                try:
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value and str(cell.value).strip():
                        samples.append(str(cell.value).strip()[:50])  # 截断长文本
                        sample_count += 1
                except:
                    continue
                    
            if sample_count >= max_samples:
                break
        
        return samples
    
    async def _intelligent_partitioning(self, worksheet, schema_direction: SchemaDirection, 
                                      table_type: TableType, regions: List[TableRegion]) -> List[Dict[str, Any]]:
        """
        智能分片
        
        参考ST-Raptor的分片策略
        """
        
        partitions = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 根据表格类型和复杂度决定分片策略
        if table_type == TableType.LIST and schema_direction == SchemaDirection.TOP:
            # 列表型表格 - 按行分片
            partitions.extend(await self._partition_by_rows(worksheet, regions))
        
        elif table_type == TableType.ATTRIBUTE and schema_direction == SchemaDirection.LEFT:
            # 属性型表格 - 按列分片
            partitions.extend(await self._partition_by_columns(worksheet, regions))
        
        elif table_type in [TableType.SEMI, TableType.MIXED]:
            # 复杂表格 - 混合分片策略
            partitions.extend(await self._partition_complex_table(worksheet, regions))
        
        else:
            # 默认分片 - 整表作为单个分片
            partitions.append({
                "partition_id": "full_table",
                "type": "complete",
                "position": (1, 1, max_row, max_col),
                "content_type": "mixed",
                "size": {"rows": max_row, "cols": max_col},
                "processing_priority": 1,
                "complexity": "medium"
            })
        
        return partitions
    
    async def _partition_by_rows(self, worksheet, regions: List[TableRegion]) -> List[Dict[str, Any]]:
        """按行分片（用于列表型表格）"""
        
        partitions = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 找到数据区域
        data_region = next((r for r in regions if r.region_type == "data"), None)
        if not data_region:
            data_region = TableRegion(1, 1, max_row, max_col, "data", 0.5)
        
        # 分片大小
        chunk_size = 20  # 每片最多20行
        
        start_row = data_region.start_row
        partition_id = 1
        
        while start_row <= data_region.end_row:
            end_row = min(start_row + chunk_size - 1, data_region.end_row)
            
            partitions.append({
                "partition_id": f"row_chunk_{partition_id}",
                "type": "row_based",
                "position": (start_row, data_region.start_col, end_row, data_region.end_col),
                "content_type": "data_rows",
                "size": {"rows": end_row - start_row + 1, "cols": data_region.end_col - data_region.start_col + 1},
                "processing_priority": partition_id,
                "complexity": "low"
            })
            
            start_row = end_row + 1
            partition_id += 1
        
        # 添加表头分片（如果存在）
        header_region = next((r for r in regions if r.region_type == "header"), None)
        if header_region:
            partitions.insert(0, {
                "partition_id": "header",
                "type": "header",
                "position": (header_region.start_row, header_region.start_col, 
                           header_region.end_row, header_region.end_col),
                "content_type": "schema",
                "size": {"rows": header_region.end_row - header_region.start_row + 1, 
                        "cols": header_region.end_col - header_region.start_col + 1},
                "processing_priority": 0,
                "complexity": "low"
            })
        
        return partitions
    
    async def _partition_by_columns(self, worksheet, regions: List[TableRegion]) -> List[Dict[str, Any]]:
        """按列分片（用于属性型表格）"""
        
        partitions = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 找到数据区域
        data_region = next((r for r in regions if r.region_type == "data"), None)
        if not data_region:
            data_region = TableRegion(1, 1, max_row, max_col, "data", 0.5)
        
        # 分片大小
        chunk_size = 10  # 每片最多10列
        
        start_col = data_region.start_col
        partition_id = 1
        
        while start_col <= data_region.end_col:
            end_col = min(start_col + chunk_size - 1, data_region.end_col)
            
            partitions.append({
                "partition_id": f"col_chunk_{partition_id}",
                "type": "column_based",
                "position": (data_region.start_row, start_col, data_region.end_row, end_col),
                "content_type": "data_columns",
                "size": {"rows": data_region.end_row - data_region.start_row + 1, 
                        "cols": end_col - start_col + 1},
                "processing_priority": partition_id,
                "complexity": "low"
            })
            
            start_col = end_col + 1
            partition_id += 1
        
        # 添加表头分片
        header_region = next((r for r in regions if r.region_type == "header"), None)
        if header_region:
            partitions.insert(0, {
                "partition_id": "header",
                "type": "header",
                "position": (header_region.start_row, header_region.start_col, 
                           header_region.end_row, header_region.end_col),
                "content_type": "schema",
                "size": {"rows": header_region.end_row - header_region.start_row + 1, 
                        "cols": header_region.end_col - header_region.start_col + 1},
                "processing_priority": 0,
                "complexity": "low"
            })
        
        return partitions
    
    async def _partition_complex_table(self, worksheet, regions: List[TableRegion]) -> List[Dict[str, Any]]:
        """复杂表格分片"""
        
        partitions = []
        
        # 为每个区域创建分片
        for i, region in enumerate(regions):
            partitions.append({
                "partition_id": f"region_{i + 1}",
                "type": "region_based",
                "position": (region.start_row, region.start_col, region.end_row, region.end_col),
                "content_type": region.region_type,
                "size": {"rows": region.end_row - region.start_row + 1, 
                        "cols": region.end_col - region.start_col + 1},
                "processing_priority": 0 if region.region_type == "header" else i,
                "complexity": "medium",
                "region_confidence": region.confidence
            })
        
        return partitions
    
    def _assess_complexity(self, worksheet, partitions: List[Dict[str, Any]], 
                          regions: List[TableRegion]) -> Dict[str, Any]:
        """评估表格复杂度"""
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        total_cells = max_row * max_col
        
        # 复杂度因子
        complexity_factors = {
            "size_factor": min(total_cells / 10000, 1.0),  # 大小因子
            "partition_factor": len(partitions) / 10,      # 分片数因子
            "merge_factor": len(worksheet.merged_cells.ranges) / total_cells if total_cells > 0 else 0,  # 合并单元格因子
            "region_factor": len(regions) / 5              # 区域数因子
        }
        
        # 综合复杂度评分
        complexity_score = sum(complexity_factors.values()) / len(complexity_factors)
        
        # 确定复杂度等级
        if complexity_score < 0.3:
            level = "low"
            strategy = "direct_processing"
        elif complexity_score < 0.7:
            level = "medium"
            strategy = "partition_processing"
        else:
            level = "high"
            strategy = "hierarchical_processing"
        
        # 生成处理建议
        recommendation = {
            "strategy": strategy,
            "parallel_processing": level in ["medium", "high"],
            "use_embedding": level == "high",
            "memory_requirement": level,
            "estimated_time": {
                "low": "< 1 minute",
                "medium": "1-3 minutes", 
                "high": "3-10 minutes"
            }[level]
        }
        
        return {
            "level": level,
            "score": complexity_score,
            "factors": complexity_factors,
            "recommendation": recommendation
        }
    
    def _region_to_dict(self, region: TableRegion) -> Dict[str, Any]:
        """将TableRegion转换为字典"""
        return {
            "start_row": region.start_row,
            "start_col": region.start_col,
            "end_row": region.end_row,
            "end_col": region.end_col,
            "region_type": region.region_type,
            "confidence": region.confidence,
            "content_sample": region.content_sample or []
        }
    
    async def _process_xls_file(self, file_path: str) -> Dict[str, Any]:
        """处理.xls文件（简化版）"""
        
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            
            result = {
                "success": True,
                "file_path": file_path,
                "sheets": [],
                "processing_strategy": {"note": "Simplified processing for XLS format"},
                "partitions": []
            }
            
            for sheet_name in workbook.sheet_names():
                worksheet = workbook.sheet_by_name(sheet_name)
                
                sheet_result = {
                    "sheet_name": sheet_name,
                    "dimensions": {"rows": worksheet.nrows, "cols": worksheet.ncols},
                    "table_type": TableType.UNKNOWN.value,
                    "complexity_level": "medium",
                    "partitions": [{
                        "partition_id": "full_sheet",
                        "type": "complete",
                        "position": (1, 1, worksheet.nrows, worksheet.ncols),
                        "content_type": "mixed",
                        "size": {"rows": worksheet.nrows, "cols": worksheet.ncols},
                        "processing_priority": 1,
                        "complexity": "medium"
                    }]
                }
                
                result["sheets"].append(sheet_result)
                
                # 添加到总分片列表
                for partition in sheet_result["partitions"]:
                    partition["source_sheet"] = sheet_name
                    result["partitions"].append(partition)
            
            return result
            
        except Exception as e:
            logger.error(f"XLS processing failed: {e}")
            return {"success": False, "error": str(e), "partitions": []}
    
    def _determine_processing_strategy(self, sheets: List[Dict[str, Any]]) -> Dict[str, Any]:
        """确定整体处理策略"""
        
        total_partitions = sum(len(sheet.get("partitions", [])) for sheet in sheets)
        complexity_levels = [sheet.get("complexity_level", "medium") for sheet in sheets]
        
        # 统计复杂度分布
        complexity_counts = {
            "low": complexity_levels.count("low"),
            "medium": complexity_levels.count("medium"),
            "high": complexity_levels.count("high")
        }
        
        # 确定策略
        if complexity_counts["high"] > 0:
            overall_strategy = "hierarchical_multimodal"
        elif complexity_counts["medium"] > complexity_counts["low"]:
            overall_strategy = "parallel_processing"
        else:
            overall_strategy = "sequential_processing"
        
        return {
            "overall_strategy": overall_strategy,
            "total_partitions": total_partitions,
            "complexity_distribution": complexity_counts,
            "recommended_parallel_agents": min(max(total_partitions // 5, 1), 4),
            "use_embedding_engine": complexity_counts["high"] > 0,
            "estimated_processing_time": f"{total_partitions * 0.5:.1f} minutes"
        }


# 导出主要的分片器实例
intelligent_partitioner = IntelligentTablePartitioner()