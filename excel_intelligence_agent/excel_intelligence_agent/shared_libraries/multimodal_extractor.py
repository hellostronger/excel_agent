"""
Multimodal Excel Content Extractor

基于ST-Raptor的多模态处理能力，实现Excel的视觉和文本内容提取。
"""

import logging
from typing import Dict, Any, List, Optional, Tuple, Union
from pathlib import Path
import tempfile
import base64
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import xlrd

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    Image = None

try:
    import matplotlib.pyplot as plt
    import matplotlib.patches as patches
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    plt = None

from .utils import setup_logging
from .embedding_engine import get_embedding_engine

# Setup logging
logger = setup_logging()


class TableStructureDetector:
    """
    表格结构检测器
    
    参考ST-Raptor的结构分析算法，检测Excel表格的层次结构。
    """
    
    def __init__(self):
        self.schema_patterns = {
            "header_keywords": ["id", "name", "date", "time", "amount", "total", "count", "value", "description", 
                               "编号", "名称", "日期", "时间", "金额", "总计", "数量", "值", "描述", "类型", "状态"],
            "summary_keywords": ["total", "sum", "average", "count", "summary", "汇总", "合计", "平均", "统计"],
            "calculation_patterns": ["=", "SUM(", "AVERAGE(", "COUNT(", "求和", "平均"]
        }
    
    def detect_schema_direction(self, worksheet) -> str:
        """
        检测表格Schema方向（顶部/左侧）
        
        Args:
            worksheet: openpyxl工作表对象
            
        Returns:
            'top' 或 'left' 或 'unknown'
        """
        max_row = min(worksheet.max_row, 10)  # 只检测前10行
        max_col = min(worksheet.max_column, 10)  # 只检测前10列
        
        # 检测顶部Schema特征
        top_score = 0
        first_row_values = []
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col)
            if cell.value:
                first_row_values.append(str(cell.value).lower())
        
        # 检查第一行是否像表头
        if first_row_values:
            header_matches = sum(1 for val in first_row_values 
                               if any(keyword in val for keyword in self.schema_patterns["header_keywords"]))
            top_score = header_matches / len(first_row_values)
        
        # 检测左侧Schema特征
        left_score = 0
        first_col_values = []
        for row in range(1, max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            if cell.value:
                first_col_values.append(str(cell.value).lower())
        
        if first_col_values:
            header_matches = sum(1 for val in first_col_values 
                               if any(keyword in val for keyword in self.schema_patterns["header_keywords"]))
            left_score = header_matches / len(first_col_values)
        
        # 判断方向
        if top_score > left_score and top_score > 0.3:
            return "top"
        elif left_score > top_score and left_score > 0.3:
            return "left"
        else:
            return "top"  # 默认为顶部
    
    def get_merge_cell_info(self, worksheet, cell_coordinate: str) -> Tuple[int, int, int, int]:
        """
        获取合并单元格信息
        
        Returns:
            (start_row, start_col, end_row, end_col)
        """
        try:
            for merged_range in worksheet.merged_cells.ranges:
                if cell_coordinate in merged_range:
                    return (merged_range.min_row, merged_range.min_col, 
                           merged_range.max_row, merged_range.max_col)
        except:
            pass
        
        # 如果不是合并单元格，返回单个单元格范围
        from openpyxl.utils import coordinate_to_tuple
        row, col = coordinate_to_tuple(cell_coordinate)
        return (row, col, row, col)
    
    def detect_table_regions(self, worksheet) -> List[Dict[str, Any]]:
        """
        检测表格中的不同区域（表头、数据、汇总等）
        
        Returns:
            区域信息列表
        """
        regions = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 检测表头区域
        schema_direction = self.detect_schema_direction(worksheet)
        
        if schema_direction == "top":
            # 寻找表头高度
            header_height = 1
            for row in range(1, min(5, max_row) + 1):
                has_content = False
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value and str(cell.value).strip():
                        has_content = True
                        break
                
                if has_content:
                    # 检查是否像表头
                    row_values = [str(worksheet.cell(row=row, column=col).value or "") 
                                 for col in range(1, max_col + 1)]
                    header_like = any(keyword in " ".join(row_values).lower() 
                                    for keyword in self.schema_patterns["header_keywords"])
                    if header_like:
                        header_height = row
                    else:
                        break
            
            regions.append({
                "type": "header",
                "range": (1, 1, header_height, max_col),
                "direction": "top"
            })
            
            if header_height < max_row:
                regions.append({
                    "type": "data",
                    "range": (header_height + 1, 1, max_row, max_col),
                    "direction": "top"
                })
        
        elif schema_direction == "left":
            # 寻找左侧表头宽度
            header_width = 1
            for col in range(1, min(5, max_col) + 1):
                has_content = False
                for row in range(1, max_row + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value and str(cell.value).strip():
                        has_content = True
                        break
                
                if has_content:
                    col_values = [str(worksheet.cell(row=row, column=col).value or "") 
                                 for row in range(1, max_row + 1)]
                    header_like = any(keyword in " ".join(col_values).lower() 
                                    for keyword in self.schema_patterns["header_keywords"])
                    if header_like:
                        header_width = col
                    else:
                        break
            
            regions.append({
                "type": "header",
                "range": (1, 1, max_row, header_width),
                "direction": "left"
            })
            
            if header_width < max_col:
                regions.append({
                    "type": "data",
                    "range": (1, header_width + 1, max_row, max_col),
                    "direction": "left"
                })
        
        return regions


class MultimodalContentExtractor:
    """
    多模态Excel内容提取器
    
    结合文本、结构和语义信息提取Excel内容。
    """
    
    def __init__(self):
        self.structure_detector = TableStructureDetector()
        self.embedding_engine = None
        
    def _get_embedding_engine(self):
        """延迟加载embedding引擎"""
        if self.embedding_engine is None:
            try:
                self.embedding_engine = get_embedding_engine()
            except Exception as e:
                logger.warning(f"Embedding engine not available: {e}")
                self.embedding_engine = None
        return self.embedding_engine
    
    async def extract_comprehensive_content(self, file_path: str) -> Dict[str, Any]:
        """
        提取Excel文件的综合内容
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            综合内容分析结果
        """
        try:
            path = Path(file_path)
            
            if path.suffix.lower() == '.xlsx':
                return await self._extract_xlsx_comprehensive(file_path)
            elif path.suffix.lower() == '.xls':
                return await self._extract_xls_comprehensive(file_path)
            else:
                raise ValueError(f"Unsupported file format: {path.suffix}")
                
        except Exception as e:
            logger.error(f"Comprehensive content extraction failed: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "content": {}
            }
    
    async def _extract_xlsx_comprehensive(self, file_path: str) -> Dict[str, Any]:
        """提取.xlsx文件的综合内容"""
        workbook = load_workbook(file_path, read_only=True, data_only=False)
        
        try:
            comprehensive_content = {
                "success": True,
                "file_format": "xlsx",
                "sheets": [],
                "global_insights": {
                    "total_sheets": len(workbook.worksheets),
                    "has_formulas": False,
                    "has_charts": False,
                    "has_images": False,
                    "complex_structures": []
                }
            }
            
            for worksheet in workbook.worksheets:
                sheet_content = await self._analyze_worksheet_comprehensive(worksheet)
                comprehensive_content["sheets"].append(sheet_content)
                
                # 更新全局信息
                if sheet_content.get("has_formulas"):
                    comprehensive_content["global_insights"]["has_formulas"] = True
                if sheet_content.get("has_charts"):
                    comprehensive_content["global_insights"]["has_charts"] = True
                if sheet_content.get("has_images"):
                    comprehensive_content["global_insights"]["has_images"] = True
            
            # 跨表关系分析
            cross_sheet_relations = await self._analyze_cross_sheet_relationships(workbook)
            comprehensive_content["cross_sheet_relations"] = cross_sheet_relations
            
            return comprehensive_content
            
        finally:
            workbook.close()
    
    async def _extract_xls_comprehensive(self, file_path: str) -> Dict[str, Any]:
        """提取.xls文件的综合内容"""
        # XLS格式的简化处理
        workbook = xlrd.open_workbook(file_path)
        
        comprehensive_content = {
            "success": True,
            "file_format": "xls",
            "sheets": [],
            "global_insights": {
                "total_sheets": len(workbook.sheet_names()),
                "has_formulas": False,
                "has_charts": False,
                "has_images": False,
                "complex_structures": []
            }
        }
        
        for sheet_name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(sheet_name)
            sheet_content = await self._analyze_xls_worksheet_comprehensive(worksheet)
            comprehensive_content["sheets"].append(sheet_content)
        
        return comprehensive_content
    
    async def _analyze_worksheet_comprehensive(self, worksheet) -> Dict[str, Any]:
        """综合分析单个工作表"""
        
        # 基础信息
        sheet_info = {
            "name": worksheet.title,
            "dimensions": {
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column
            },
            "structure_analysis": {},
            "content_analysis": {},
            "semantic_analysis": {},
            "visual_elements": {},
            "has_formulas": False,
            "has_charts": False,
            "has_images": False
        }
        
        # 结构分析
        try:
            regions = self.structure_detector.detect_table_regions(worksheet)
            schema_direction = self.structure_detector.detect_schema_direction(worksheet)
            
            sheet_info["structure_analysis"] = {
                "schema_direction": schema_direction,
                "detected_regions": regions,
                "merge_cells_count": len(worksheet.merged_cells.ranges),
                "table_structure": self._classify_table_structure(worksheet, regions)
            }
        except Exception as e:
            logger.error(f"Structure analysis failed: {e}")
            sheet_info["structure_analysis"] = {"error": str(e)}
        
        # 内容分析
        try:
            content_analysis = await self._analyze_worksheet_content(worksheet)
            sheet_info["content_analysis"] = content_analysis
            
            # 检测公式
            sheet_info["has_formulas"] = content_analysis.get("formula_count", 0) > 0
        except Exception as e:
            logger.error(f"Content analysis failed: {e}")
            sheet_info["content_analysis"] = {"error": str(e)}
        
        # 语义分析
        try:
            if self._get_embedding_engine():
                semantic_analysis = await self._analyze_worksheet_semantics(worksheet)
                sheet_info["semantic_analysis"] = semantic_analysis
        except Exception as e:
            logger.error(f"Semantic analysis failed: {e}")
            sheet_info["semantic_analysis"] = {"error": str(e)}
        
        # 视觉元素检测
        try:
            visual_elements = await self._detect_visual_elements(worksheet)
            sheet_info["visual_elements"] = visual_elements
            sheet_info["has_charts"] = len(visual_elements.get("charts", [])) > 0
            sheet_info["has_images"] = len(visual_elements.get("images", [])) > 0
        except Exception as e:
            logger.error(f"Visual elements detection failed: {e}")
            sheet_info["visual_elements"] = {"error": str(e)}
        
        return sheet_info
    
    def _classify_table_structure(self, worksheet, regions: List[Dict[str, Any]]) -> str:
        """
        分类表格结构类型
        
        参考ST-Raptor的分类方法：
        - T_LIST: 列表型表格
        - T_ATTR: 属性型表格  
        - T_SEMI: 半结构化表格
        - T_MIX: 混合型表格
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # 简单的结构分类逻辑
        if max_row <= 3 and max_col <= 3:
            return "simple"
        elif len(regions) == 1:
            return "single_table"
        elif any(region["type"] == "header" for region in regions):
            if any(region["direction"] == "top" for region in regions):
                return "list_type"  # 类似T_LIST
            else:
                return "attribute_type"  # 类似T_ATTR
        else:
            return "mixed_type"  # 类似T_MIX
    
    async def _analyze_worksheet_content(self, worksheet) -> Dict[str, Any]:
        """分析工作表内容"""
        
        content_stats = {
            "total_cells": worksheet.max_row * worksheet.max_column,
            "non_empty_cells": 0,
            "formula_count": 0,
            "numeric_cells": 0,
            "text_cells": 0,
            "date_cells": 0,
            "data_types": {},
            "cell_samples": []
        }
        
        # 采样分析（避免处理超大表格）
        sample_size = min(1000, worksheet.max_row * worksheet.max_column)
        cells_analyzed = 0
        
        for row in worksheet.iter_rows(max_row=min(worksheet.max_row, 50)):
            if cells_analyzed >= sample_size:
                break
                
            for cell in row:
                if cells_analyzed >= sample_size:
                    break
                    
                cells_analyzed += 1
                
                if cell.value is not None:
                    content_stats["non_empty_cells"] += 1
                    
                    # 检测数据类型
                    cell_value = cell.value
                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        content_stats["formula_count"] += 1
                    elif isinstance(cell_value, (int, float)):
                        content_stats["numeric_cells"] += 1
                    elif isinstance(cell_value, str):
                        content_stats["text_cells"] += 1
                    
                    # 收集样本
                    if len(content_stats["cell_samples"]) < 20:
                        content_stats["cell_samples"].append({
                            "coordinate": cell.coordinate,
                            "value": str(cell_value)[:100],  # 截断长文本
                            "data_type": type(cell_value).__name__
                        })
        
        # 计算密度
        if content_stats["total_cells"] > 0:
            content_stats["data_density"] = content_stats["non_empty_cells"] / content_stats["total_cells"]
        else:
            content_stats["data_density"] = 0.0
        
        return content_stats
    
    async def _analyze_worksheet_semantics(self, worksheet) -> Dict[str, Any]:
        """分析工作表的语义特征"""
        
        engine = self._get_embedding_engine()
        if not engine:
            return {"error": "Embedding engine not available"}
        
        semantic_info = {
            "key_terms": [],
            "content_categories": {},
            "column_semantics": {},
            "semantic_coherence": 0.0
        }
        
        try:
            # 提取文本内容用于语义分析
            text_content = []
            column_contents = {}
            
            # 分析列内容
            for col in range(1, min(worksheet.max_column, 20) + 1):  # 限制列数
                col_values = []
                for row in range(1, min(worksheet.max_row, 100) + 1):  # 限制行数
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str) and len(cell.value.strip()) > 0:
                        col_values.append(cell.value.strip())
                        text_content.append(cell.value.strip())
                
                if col_values:
                    # 为列生成语义特征
                    column_letter = chr(ord('A') + col - 1) if col <= 26 else f"Column_{col}"
                    try:
                        col_semantics = await self._analyze_column_semantic_coherence(col_values)
                        semantic_info["column_semantics"][column_letter] = col_semantics
                    except Exception as e:
                        logger.error(f"Column semantic analysis failed: {e}")
            
            # 整体语义分析
            if text_content:
                # 提取关键术语（高频词汇）
                from collections import Counter
                words = []
                for text in text_content:
                    # 简单的词汇提取
                    text_words = [word.strip() for word in text.replace(',', ' ').replace(';', ' ').split() 
                                 if len(word.strip()) > 2]
                    words.extend(text_words)
                
                if words:
                    word_freq = Counter(words)
                    semantic_info["key_terms"] = [word for word, freq in word_freq.most_common(10)]
            
            return semantic_info
            
        except Exception as e:
            logger.error(f"Semantic analysis failed: {e}")
            return {"error": str(e)}
    
    async def _analyze_column_semantic_coherence(self, column_values: List[str]) -> Dict[str, Any]:
        """分析列的语义一致性"""
        
        engine = self._get_embedding_engine()
        if not engine or not column_values:
            return {"coherence_score": 0.0, "categories": []}
        
        try:
            # 计算语义相似度
            if len(column_values) > 1:
                embeddings = engine.encode_entities(column_values)
                
                # 计算平均成对相似度
                from sklearn.metrics.pairwise import cosine_similarity
                similarity_matrix = cosine_similarity(embeddings)
                
                # 排除对角线
                import numpy as np
                mask = np.ones(similarity_matrix.shape, dtype=bool)
                np.fill_diagonal(mask, False)
                coherence_score = similarity_matrix[mask].mean()
            else:
                coherence_score = 1.0
            
            return {
                "coherence_score": float(coherence_score),
                "unique_values": len(set(column_values)),
                "total_values": len(column_values),
                "sample_values": column_values[:5]
            }
            
        except Exception as e:
            logger.error(f"Column coherence analysis failed: {e}")
            return {"coherence_score": 0.0, "error": str(e)}
    
    async def _detect_visual_elements(self, worksheet) -> Dict[str, Any]:
        """检测视觉元素（图表、图片等）"""
        
        visual_elements = {
            "charts": [],
            "images": [],
            "shapes": [],
            "conditional_formatting": []
        }
        
        try:
            # 检测图表
            if hasattr(worksheet, '_charts'):
                for chart in worksheet._charts:
                    visual_elements["charts"].append({
                        "type": type(chart).__name__,
                        "title": getattr(chart, 'title', None),
                        "anchor": str(chart.anchor) if hasattr(chart, 'anchor') else None
                    })
            
            # 检测图片
            if hasattr(worksheet, '_images'):
                for img in worksheet._images:
                    visual_elements["images"].append({
                        "format": getattr(img, 'format', None),
                        "anchor": str(img.anchor) if hasattr(img, 'anchor') else None
                    })
            
            # 检测条件格式
            if hasattr(worksheet, 'conditional_formatting'):
                for cf in worksheet.conditional_formatting:
                    visual_elements["conditional_formatting"].append({
                        "range": str(cf.sqref) if hasattr(cf, 'sqref') else None,
                        "type": str(type(cf).__name__)
                    })
                    
        except Exception as e:
            logger.error(f"Visual elements detection failed: {e}")
            visual_elements["error"] = str(e)
        
        return visual_elements
    
    async def _analyze_xls_worksheet_comprehensive(self, worksheet) -> Dict[str, Any]:
        """分析XLS工作表（简化版）"""
        
        sheet_info = {
            "name": worksheet.name,
            "dimensions": {
                "max_row": worksheet.nrows,
                "max_column": worksheet.ncols
            },
            "structure_analysis": {"note": "Limited analysis for XLS format"},
            "content_analysis": {},
            "semantic_analysis": {"note": "Limited semantic analysis for XLS format"},
            "visual_elements": {"note": "Visual elements detection not supported for XLS"},
            "has_formulas": False,
            "has_charts": False,
            "has_images": False
        }
        
        # 基础内容分析
        try:
            content_stats = {
                "total_cells": worksheet.nrows * worksheet.ncols,
                "non_empty_cells": 0,
                "formula_count": 0,
                "cell_samples": []
            }
            
            # 采样分析
            for row_idx in range(min(worksheet.nrows, 50)):
                for col_idx in range(min(worksheet.ncols, 20)):
                    try:
                        cell = worksheet.cell(row_idx, col_idx)
                        if cell.value is not None and cell.value != '':
                            content_stats["non_empty_cells"] += 1
                            
                            if cell.ctype == xlrd.XL_CELL_FORMULA:
                                content_stats["formula_count"] += 1
                                sheet_info["has_formulas"] = True
                            
                            if len(content_stats["cell_samples"]) < 10:
                                content_stats["cell_samples"].append({
                                    "coordinate": f"R{row_idx+1}C{col_idx+1}",
                                    "value": str(cell.value)[:50],
                                    "cell_type": cell.ctype
                                })
                    except:
                        continue
            
            if content_stats["total_cells"] > 0:
                content_stats["data_density"] = content_stats["non_empty_cells"] / content_stats["total_cells"]
            else:
                content_stats["data_density"] = 0.0
            
            sheet_info["content_analysis"] = content_stats
            
        except Exception as e:
            logger.error(f"XLS content analysis failed: {e}")
            sheet_info["content_analysis"] = {"error": str(e)}
        
        return sheet_info
    
    async def _analyze_cross_sheet_relationships(self, workbook) -> Dict[str, Any]:
        """分析跨表关系"""
        
        relationships = {
            "formula_references": [],
            "data_flow": [],
            "shared_headers": [],
            "potential_links": []
        }
        
        try:
            sheet_names = [ws.title for ws in workbook.worksheets]
            
            # 检测公式引用
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(max_row=min(worksheet.max_row, 20)):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula = cell.value
                            # 查找跨表引用
                            for other_sheet in sheet_names:
                                if other_sheet != worksheet.title and other_sheet in formula:
                                    relationships["formula_references"].append({
                                        "source_sheet": worksheet.title,
                                        "target_sheet": other_sheet,
                                        "cell": cell.coordinate,
                                        "formula": formula[:100]
                                    })
            
            # 如果有embedding引擎，分析共享列名
            engine = self._get_embedding_engine()
            if engine:
                sheet_headers = {}
                for worksheet in workbook.worksheets:
                    headers = []
                    # 提取可能的表头
                    for col in range(1, min(worksheet.max_column, 10) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        if cell.value and isinstance(cell.value, str):
                            headers.append(cell.value.strip())
                    sheet_headers[worksheet.title] = headers
                
                # 找到相似的表头
                for sheet1, headers1 in sheet_headers.items():
                    for sheet2, headers2 in sheet_headers.items():
                        if sheet1 != sheet2 and headers1 and headers2:
                            # 使用embedding计算相似度
                            similar_pairs = []
                            for h1 in headers1:
                                similar = engine.semantic_similarity(h1, headers2, top_k=1)
                                if similar and similar[0][1] > 0.7:  # 相似度阈值
                                    similar_pairs.append((h1, similar[0][0], similar[0][1]))
                            
                            if similar_pairs:
                                relationships["shared_headers"].append({
                                    "sheet1": sheet1,
                                    "sheet2": sheet2,
                                    "similar_headers": similar_pairs
                                })
        
        except Exception as e:
            logger.error(f"Cross-sheet analysis failed: {e}")
            relationships["error"] = str(e)
        
        return relationships


# 导出主要的提取器实例
multimodal_extractor = MultimodalContentExtractor()