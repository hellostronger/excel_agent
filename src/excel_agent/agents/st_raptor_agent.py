"""ST-Raptor Enhanced Agent for advanced Excel processing with hierarchical feature trees."""

import os
import json
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from .base import BaseAgent
from ..models.base import AgentRequest, AgentResponse, FileMetadata
from ..models.feature_tree import FeatureTree, IndexNode, TreeNode
from ..utils.config import get_config
from ..utils.cache_manager import get_cache_manager
from .embedding_agent import EmbeddingAgent


class STRaptorAgent(BaseAgent):
    """ST-Raptor enhanced agent for advanced Excel processing."""
    
    def __init__(self):
        super().__init__("st_raptor_agent")
        self.config = get_config()
        self.cache_manager = get_cache_manager()
        self.embedding_agent = EmbeddingAgent()
        self.processed_files = {}  # Store for ST-Raptor processed files
        
    async def process(self, request: AgentRequest) -> AgentResponse:
        """Process file with ST-Raptor enhancements."""
        try:
            file_path = request.data.get("file_path")
            if not file_path:
                return AgentResponse(
                    request_id=request.request_id,
                    agent_name=self.name,
                    status="error",
                    error_message="Missing file_path in request"
                )
            
            path = Path(file_path)
            if not path.exists():
                return AgentResponse(
                    request_id=request.request_id,
                    agent_name=self.name,
                    status="error",
                    error_message=f"File not found: {file_path}"
                )
            
            # Generate file ID
            file_id = self._generate_file_id(path)
            
            # Check cache first
            if self.config.enable_cache:
                cached_result = self._load_from_cache(file_id)
                if cached_result:
                    self.logger.info(f"Loaded ST-Raptor data from cache: {file_id}")
                    return cached_result
            
            # Process with ST-Raptor enhancements
            result = await self._process_with_enhancements(path, file_id)
            
            # Save to cache
            if self.config.enable_cache and result.status == "completed":
                self._save_to_cache(file_id, result)
            
            return result
            
        except Exception as e:
            self.logger.error(f"Error in ST-Raptor processing: {e}")
            return AgentResponse(
                request_id=request.request_id,
                agent_name=self.name,
                status="error",
                error_message=str(e)
            )
    
    async def _process_with_enhancements(self, file_path: Path, file_id: str) -> AgentResponse:
        """Process file with all ST-Raptor enhancements."""
        try:
            # Step 1: Parse file metadata
            metadata = await self._parse_excel_metadata(file_path, file_id)
            
            # Step 2: Create hierarchical feature tree
            feature_tree = await self._create_feature_tree(file_path, file_id, metadata)
            
            # Step 3: Generate embeddings for semantic search
            embedding_dict = None
            if self.config.enable_embedding_cache:
                embedding_dict = self.embedding_agent.create_embedding_dict(feature_tree)
                self.cache_manager.save_embedding_cache(embedding_dict, file_id)
            
            # Step 4: Store enhanced data
            enhanced_data = {
                "metadata": metadata,
                "feature_tree": feature_tree,
                "embedding_dict": embedding_dict,
                "processing_mode": "st_raptor",
                "processed_timestamp": datetime.now().isoformat()
            }
            
            self.processed_files[file_id] = enhanced_data
            
            self.logger.info(f"Successfully processed file with ST-Raptor: {file_path.name}")
            
            # Return comprehensive response
            return AgentResponse(
                request_id=file_id,
                agent_name=self.name,
                status="completed",
                data={
                    "file_id": file_id,
                    "file_path": str(file_path),
                    "processing_mode": "st_raptor",
                    "metadata": metadata.__dict__ if hasattr(metadata, '__dict__') else metadata,
                    "feature_tree_stats": feature_tree.get_statistics(),
                    "embedding_stats": {
                        "count": embedding_dict.get("count", 0) if embedding_dict else 0,
                        "dimension": embedding_dict.get("dimension", 0) if embedding_dict else 0,
                        "model": embedding_dict.get("model_name", "unknown") if embedding_dict else "none"
                    },
                    "enhancements": [
                        "hierarchical_feature_trees",
                        "semantic_embeddings", 
                        "intelligent_caching",
                        "structure_analysis"
                    ]
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error in enhanced processing: {e}")
            return AgentResponse(
                request_id=file_id,
                agent_name=self.name,
                status="error",
                error_message=str(e)
            )
    
    async def _parse_excel_metadata(self, file_path: Path, file_id: str) -> FileMetadata:
        """Parse Excel file and extract enhanced metadata."""
        try:
            workbook = openpyxl.load_workbook(
                file_path,
                read_only=False,
                data_only=False
            )
            
            sheets = []
            total_rows = 0
            total_columns = 0
            has_merged_cells = False
            has_formulas = False
            has_charts = False
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheets.append(sheet_name)
                
                if sheet.max_row:
                    total_rows += sheet.max_row
                if sheet.max_column:
                    total_columns += sheet.max_column
                
                if sheet.merged_cells:
                    has_merged_cells = True
                
                # Check for formulas
                for row in sheet.iter_rows(max_row=min(100, sheet.max_row or 100)):
                    for cell in row:
                        if cell.data_type == 'f':
                            has_formulas = True
                            break
                    if has_formulas:
                        break
                
                # Check for charts
                if hasattr(sheet, '_charts') and sheet._charts:
                    has_charts = True
            
            workbook.close()
            
            return FileMetadata(
                file_id=file_id,
                original_filename=file_path.name,
                file_path=str(file_path.absolute()),
                file_size=file_path.stat().st_size,
                upload_timestamp=datetime.now(),
                sheets=sheets,
                total_rows=total_rows,
                total_columns=total_columns,
                has_merged_cells=has_merged_cells,
                has_formulas=has_formulas,
                has_charts=has_charts,
                has_images=False  # Could be enhanced further
            )
            
        except Exception as e:
            self.logger.error(f"Error parsing Excel metadata: {e}")
            raise
    
    async def _create_feature_tree(self, file_path: Path, file_id: str, metadata: FileMetadata) -> FeatureTree:
        """Create hierarchical feature tree (ST-Raptor inspired)."""
        try:
            tree = FeatureTree()
            tree.table_id = file_id
            tree.file_path = str(file_path)
            tree.set_metadata(metadata.__dict__)
            
            # Create root node
            root = IndexNode(value=f"Excel_File_{file_path.stem}")
            
            # Process sheets
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in metadata.sheets:
                sheet = workbook[sheet_name]
                sheet_node = IndexNode(value=f"Sheet_{sheet_name}")
                
                # Add columns with data sampling
                if sheet.max_column and sheet.max_column > 0:
                    max_cols = min(sheet.max_column, 50)  # Limit for performance
                    max_rows = min(sheet.max_row or 100, 100)
                    
                    for col_idx in range(1, max_cols + 1):
                        col_letter = get_column_letter(col_idx)
                        col_node = IndexNode(value=f"Column_{col_letter}")
                        
                        # Sample cell values
                        cell_values = []
                        for row_idx in range(1, max_rows + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None:
                                cell_values.append(TreeNode(value=str(cell.value)[:100]))  # Truncate long values
                        
                        # Add sample values to column node
                        for cell_node in cell_values[:20]:  # Limit to 20 samples per column
                            col_node.add_body_node(cell_node)
                        
                        sheet_node.add_child(col_node)
                
                # Handle merged cells
                if sheet.merged_cells:
                    merged_node = IndexNode(value="Merged_Cells")
                    for merged_range in list(sheet.merged_cells)[:10]:
                        range_node = TreeNode(value=str(merged_range))
                        merged_node.add_body_node(range_node)
                    sheet_node.add_child(merged_node)
                
                root.add_child(sheet_node)
            
            workbook.close()
            tree.set_root(root)
            
            # Set schema information
            schema_info = {
                "file_type": "excel",
                "sheets": metadata.sheets,
                "complexity": self._assess_structure_complexity(metadata),
                "features": {
                    "merged_cells": metadata.has_merged_cells,
                    "formulas": metadata.has_formulas,
                    "charts": metadata.has_charts
                }
            }
            tree.set_schema_info(schema_info)
            
            return tree
            
        except Exception as e:
            self.logger.error(f"Error creating feature tree: {e}")
            # Return minimal tree as fallback
            tree = FeatureTree()
            tree.table_id = file_id
            tree.set_root(IndexNode(value="Error_Processing"))
            return tree
    
    def _assess_structure_complexity(self, metadata: FileMetadata) -> str:
        """Assess structural complexity for processing optimization."""
        complexity_score = 0
        
        if len(metadata.sheets) > 1:
            complexity_score += 1
        if metadata.total_rows > 1000:
            complexity_score += 1
        if metadata.total_columns > 50:
            complexity_score += 1
        if metadata.has_merged_cells:
            complexity_score += 1
        if metadata.has_formulas:
            complexity_score += 1
        if metadata.has_charts:
            complexity_score += 1
        
        if complexity_score >= 4:
            return "high"
        elif complexity_score >= 2:
            return "medium"
        else:
            return "low"
    
    def _generate_file_id(self, file_path: Path) -> str:
        """Generate unique file ID."""
        file_stat = file_path.stat()
        content = f"st_raptor:{file_path.absolute()}:{file_stat.st_mtime}:{file_stat.st_size}"
        return hashlib.md5(content.encode()).hexdigest()
    
    def _load_from_cache(self, file_id: str) -> Optional[AgentResponse]:
        """Load ST-Raptor processed data from cache."""
        try:
            # Load metadata
            metadata_cache = self.cache_manager.load_metadata_cache(file_id)
            if not metadata_cache:
                return None
            
            # Load feature tree
            feature_tree = self.cache_manager.load_tree_cache(file_id)
            if not feature_tree:
                return None
            
            # Load embeddings
            embedding_dict = self.cache_manager.load_embedding_cache(file_id)
            
            # Restore data
            enhanced_data = {
                "metadata": metadata_cache,
                "feature_tree": feature_tree,
                "embedding_dict": embedding_dict,
                "processing_mode": "st_raptor"
            }
            
            self.processed_files[file_id] = enhanced_data
            
            return AgentResponse(
                request_id=file_id,
                agent_name=self.name,
                status="completed",
                data={
                    "file_id": file_id,
                    "processing_mode": "st_raptor",
                    "metadata": metadata_cache,
                    "feature_tree_stats": feature_tree.get_statistics(),
                    "embedding_stats": {
                        "count": embedding_dict.get("count", 0) if embedding_dict else 0,
                        "dimension": embedding_dict.get("dimension", 0) if embedding_dict else 0
                    },
                    "cached": True
                }
            )
            
        except Exception as e:
            self.logger.warning(f"Failed to load from cache: {e}")
            return None
    
    def _save_to_cache(self, file_id: str, response: AgentResponse):
        """Save ST-Raptor data to cache."""
        try:
            if file_id not in self.processed_files:
                return
            
            data = self.processed_files[file_id]
            
            # Save components
            if "metadata" in data:
                self.cache_manager.save_metadata_cache(data["metadata"].__dict__, file_id)
            
            if "feature_tree" in data:
                self.cache_manager.save_tree_cache(data["feature_tree"], file_id)
            
            # Embedding is already saved during processing
            
        except Exception as e:
            self.logger.warning(f"Failed to save to cache: {e}")
    
    def get_processed_file(self, file_id: str) -> Optional[Dict[str, Any]]:
        """Get processed file data for other agents."""
        return self.processed_files.get(file_id)
    
    def get_feature_tree(self, file_id: str) -> Optional[FeatureTree]:
        """Get feature tree for a processed file."""
        data = self.processed_files.get(file_id)
        if data:
            return data.get("feature_tree")
        return None
    
    def get_embedding_dict(self, file_id: str) -> Optional[Dict[str, Any]]:
        """Get embedding dictionary for a processed file."""
        data = self.processed_files.get(file_id)
        if data:
            return data.get("embedding_dict")
        return None