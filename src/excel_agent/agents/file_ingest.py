"""File Ingest Agent for loading and parsing Excel files."""

import os
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import openpyxl

from .base import BaseAgent
from ..models.agents import FileIngestRequest, FileIngestResponse
from ..models.base import FileMetadata, AgentRequest, AgentResponse, AgentStatus
from ..utils.config import config


class FileIngestAgent(BaseAgent):
    """Agent responsible for loading and parsing Excel files."""
    
    def __init__(self):
        super().__init__(
            name="FileIngestAgent",
            description="Loads and parses Excel files, extracts sheet names and metadata",
            mcp_capabilities=["excel_tools", "file_management"]
        )
        self.file_store = {}  # In-memory store for file metadata
    
    async def process(self, request: AgentRequest) -> AgentResponse:
        """Process file ingestion request."""
        if not isinstance(request, FileIngestRequest):
            return self.create_error_response(
                request, 
                f"Invalid request type. Expected FileIngestRequest, got {type(request)}"
            )
        
        try:
            # Validate file path
            file_path = Path(request.file_path)
            if not file_path.exists():
                return self.create_error_response(
                    request,
                    f"File does not exist: {request.file_path}"
                )
            
            if not file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
                return self.create_error_response(
                    request,
                    f"Unsupported file format: {file_path.suffix}"
                )
            
            # Check file size
            file_size = file_path.stat().st_size
            if file_size > config.max_file_size_mb * 1024 * 1024:
                return self.create_error_response(
                    request,
                    f"File too large: {file_size / (1024*1024):.1f}MB > {config.max_file_size_mb}MB"
                )
            
            # Generate file ID
            file_id = self._generate_file_id(request.file_path)
            
            # Parse Excel file
            metadata = await self._parse_excel_file(file_path, file_id)
            
            # Use MCP tools for enhanced file processing if available
            mcp_result = await self._enhance_with_mcp(file_path, metadata)
            if mcp_result:
                metadata = mcp_result
            
            # Store metadata
            self.file_store[file_id] = metadata
            
            self.logger.info(f"Successfully ingested file: {file_path.name}")
            
            # Create response
            return FileIngestResponse(
                agent_id=self.name,
                request_id=request.request_id,
                status=AgentStatus.SUCCESS,
                file_id=file_id,
                sheets=metadata.sheets,
                result={
                    "file_id": file_id,
                    "sheets": metadata.sheets,
                    "metadata": metadata.model_dump()
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error processing file ingestion: {e}")
            return self.create_error_response(request, str(e))
    
    async def _parse_excel_file(self, file_path: Path, file_id: str) -> FileMetadata:
        """Parse Excel file and extract metadata."""
        try:
            # Load with openpyxl for rich metadata
            workbook = openpyxl.load_workbook(
                file_path, 
                read_only=False,
                keep_vba=True,
                data_only=False,
                keep_links=True
            )
            
            sheets = []
            total_rows = 0
            total_columns = 0
            has_merged_cells = False
            has_formulas = False
            has_charts = False
            has_images = False
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheets.append(sheet_name)
                
                # Count rows and columns
                if sheet.max_row:
                    total_rows += sheet.max_row
                if sheet.max_column:
                    total_columns += sheet.max_column
                
                # Check for merged cells
                if sheet.merged_cells:
                    has_merged_cells = True
                
                # Check for formulas
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            has_formulas = True
                            break
                    if has_formulas:
                        break
                
                # Check for charts
                if hasattr(sheet, '_charts') and sheet._charts:
                    has_charts = True
                
                # Check for images
                if hasattr(sheet, '_images') and sheet._images:
                    has_images = True
            
            workbook.close()
            
            # Create metadata
            metadata = FileMetadata(
                file_id=file_id,
                original_filename=file_path.name,
                file_path=str(file_path.absolute()),
                file_size=file_path.stat().st_size,
                upload_timestamp=datetime.now(),
                sheets=sheets,
                total_rows=total_rows,
                total_columns=total_columns,
                has_merged_cells=has_merged_cells,
                has_formulas=has_formulas,
                has_charts=has_charts,
                has_images=has_images
            )
            
            return metadata
            
        except Exception as e:
            # Fallback to pandas for basic info
            self.logger.warning(f"openpyxl failed, falling back to pandas: {e}")
            
            try:
                # Read all sheets with pandas
                excel_file = pd.ExcelFile(file_path)
                sheets = excel_file.sheet_names
                
                total_rows = 0
                total_columns = 0
                
                for sheet_name in sheets:
                    df = excel_file.parse(sheet_name)
                    total_rows += len(df)
                    total_columns += len(df.columns)
                
                metadata = FileMetadata(
                    file_id=file_id,
                    original_filename=file_path.name,
                    file_path=str(file_path.absolute()),
                    file_size=file_path.stat().st_size,
                    upload_timestamp=datetime.now(),
                    sheets=sheets,
                    total_rows=total_rows,
                    total_columns=total_columns,
                    has_merged_cells=False,  # Unknown with pandas
                    has_formulas=False,     # Unknown with pandas
                    has_charts=False,       # Unknown with pandas
                    has_images=False        # Unknown with pandas
                )
                
                return metadata
                
            except Exception as e2:
                raise Exception(f"Failed to parse Excel file with both openpyxl and pandas: {e2}")
    
    def _generate_file_id(self, file_path: str) -> str:
        """Generate a unique file ID based on file path and timestamp."""
        timestamp = datetime.now().isoformat()
        content = f"{file_path}:{timestamp}"
        return hashlib.md5(content.encode()).hexdigest()
    
    def get_file_metadata(self, file_id: str) -> Optional[FileMetadata]:
        """Get file metadata by file ID."""
        return self.file_store.get(file_id)
    
    def list_files(self) -> List[FileMetadata]:
        """List all ingested files."""
        return list(self.file_store.values())
    
    def remove_file(self, file_id: str) -> bool:
        """Remove file from store."""
        if file_id in self.file_store:
            del self.file_store[file_id]
            return True
        return False
    
    async def _enhance_with_mcp(self, file_path: Path, metadata: FileMetadata) -> Optional[FileMetadata]:
        """Enhance file metadata using MCP tools."""
        try:
            # Use MCP Excel tools for enhanced analysis
            excel_info = await self.call_mcp_tool("read_excel_file", {
                "file_path": str(file_path)
            })
            
            if excel_info and not excel_info.get("error"):
                # Create backup using MCP file management
                backup_result = await self.call_mcp_tool("create_backup", {
                    "file_path": str(file_path)
                })
                
                # Enhance metadata with MCP results
                enhanced_metadata = metadata.model_copy()
                
                if excel_info.get("sheets"):
                    enhanced_metadata.sheets = list(excel_info["sheets"].keys())
                    
                    # Add sheet-specific information
                    total_rows = 0
                    total_columns = 0
                    for sheet_info in excel_info["sheets"].values():
                        if "shape" in sheet_info:
                            total_rows += sheet_info["shape"][0]
                            total_columns += sheet_info["shape"][1]
                    
                    enhanced_metadata.total_rows = total_rows
                    enhanced_metadata.total_columns = total_columns
                
                # Log backup status
                if backup_result and not backup_result.get("error"):
                    self.logger.info(f"Created backup: {backup_result.get('backup_file')}")
                
                return enhanced_metadata
            
            return None
            
        except Exception as e:
            self.logger.warning(f"MCP enhancement failed: {e}")
            return None