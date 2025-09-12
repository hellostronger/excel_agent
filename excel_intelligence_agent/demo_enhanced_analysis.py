#!/usr/bin/env python3
"""
Enhanced Excel Intelligence Agent Demo

演示集成了ST-Raptor技术的增强Excel分析功能：
- 多模态内容提取
- Embedding驱动的语义分析
- 智能表格分片
- 树状结构理解
"""

import os
import sys
import asyncio
import tempfile
from pathlib import Path
import json

# 添加项目路径
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

def setup_environment():
    """设置环境变量"""
    env_file = project_root / ".env"
    if env_file.exists():
        with open(env_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key.strip()] = value.strip()


async def demo_enhanced_analysis():
    """演示增强的Excel分析功能"""
    
    print("=== Enhanced Excel Intelligence Agent Demo ===")
    print("集成ST-Raptor技术的增强Excel分析系统\n")
    
    try:
        # 导入增强的模块
        from excel_intelligence_agent.shared_libraries.embedding_engine import get_embedding_engine
        from excel_intelligence_agent.shared_libraries.multimodal_extractor import multimodal_extractor
        from excel_intelligence_agent.shared_libraries.tree_partitioner import intelligent_partitioner
        from excel_intelligence_agent.sub_agents.file_analyzer.agent import file_analyzer_agent
        
        print("✓ 成功导入增强分析模块")
        
        # 测试embedding引擎
        print("\n=== 测试Embedding引擎 ===")
        try:
            embedding_engine = get_embedding_engine()
            print("✓ Embedding引擎初始化成功")
            
            # 语义相似度测试
            test_queries = ["财务报表", "销售数据", "员工信息"]
            test_candidates = ["收入支出表", "业绩统计", "人力资源", "库存管理", "客户列表"]
            
            print("测试语义相似度匹配...")
            for query in test_queries:
                results = embedding_engine.semantic_similarity(query, test_candidates, top_k=2)
                print(f"查询'{query}': {results}")
                
        except Exception as e:
            print(f"⚠ Embedding引擎测试失败: {e}")
            print("注意: 这是正常的，需要安装sentence-transformers库")
        
        # 测试智能分片器
        print("\n=== 测试智能分片器 ===")
        try:
            # 创建一个简单的测试Excel文件
            test_file = await create_test_excel()
            print(f"✓ 创建测试文件: {test_file}")
            
            # 进行智能分片分析
            partition_result = await intelligent_partitioner.analyze_and_partition(test_file)
            
            if partition_result.get("success"):
                print("✓ 智能分片分析成功")
                print(f"  - 检测到 {len(partition_result.get('sheets', []))} 个工作表")
                print(f"  - 生成 {len(partition_result.get('partitions', []))} 个分片")
                print(f"  - 处理策略: {partition_result.get('processing_strategy', {}).get('overall_strategy', 'unknown')}")
            else:
                print(f"⚠ 智能分片分析失败: {partition_result.get('error')}")
                
            # 清理测试文件
            try:
                os.unlink(test_file)
            except:
                pass
                
        except Exception as e:
            print(f"⚠ 智能分片器测试失败: {e}")
        
        # 测试多模态提取器
        print("\n=== 测试多模态内容提取 ===")
        try:
            test_file = await create_test_excel()
            
            multimodal_result = await multimodal_extractor.extract_comprehensive_content(test_file)
            
            if multimodal_result.get("success"):
                print("✓ 多模态内容提取成功")
                sheets = multimodal_result.get("sheets", [])
                print(f"  - 分析了 {len(sheets)} 个工作表")
                
                for sheet in sheets[:2]:  # 只显示前2个
                    print(f"  - 工作表'{sheet.get('name')}': {sheet.get('dimensions', {})}")
                    structure = sheet.get("structure_analysis", {})
                    print(f"    Schema方向: {structure.get('schema_direction', 'unknown')}")
            else:
                print(f"⚠ 多模态提取失败: {multimodal_result.get('error')}")
                
            # 清理
            try:
                os.unlink(test_file)
            except:
                pass
                
        except Exception as e:
            print(f"⚠ 多模态提取器测试失败: {e}")
        
        # 测试增强的文件分析器
        print("\n=== 测试增强文件分析器 ===")
        try:
            test_file = await create_complex_test_excel()
            print(f"✓ 创建复杂测试文件: {test_file}")
            
            # 模拟使用增强的文件分析器
            from excel_intelligence_agent.sub_agents.file_analyzer.tools import analyze_file_structure
            from excel_intelligence_agent.shared_libraries.utils import extract_basic_metadata
            from google.adk.tools import ToolContext
            
            # 创建工具上下文
            tool_context = ToolContext()
            metadata = await extract_basic_metadata(test_file)
            tool_context.state["file_metadata"] = metadata.dict()
            
            # 执行增强分析
            analysis_result = await analyze_file_structure(test_file, "comprehensive", tool_context)
            
            if analysis_result.get("success"):
                print("✓ 增强文件分析成功")
                print(f"  - 基础信息: {analysis_result.get('sheet_count', 0)} 个工作表")
                print(f"  - 数据单元格: {analysis_result.get('total_data_cells', 0)} 个")
                
                # 显示增强功能
                if "multimodal_analysis" in analysis_result:
                    ma_success = analysis_result["multimodal_analysis"].get("success", False)
                    print(f"  - 多模态分析: {'✓' if ma_success else '⚠'}")
                
                if "partitioning_analysis" in analysis_result:
                    pa_success = analysis_result["partitioning_analysis"].get("success", False)
                    print(f"  - 分片分析: {'✓' if pa_success else '⚠'}")
                
                if "semantic_analysis" in analysis_result:
                    sa_error = analysis_result["semantic_analysis"].get("error")
                    print(f"  - 语义分析: {'✓' if not sa_error else '⚠'}")
                
                # 显示增强建议
                enhanced_recommendations = analysis_result.get("enhanced_recommendations", [])
                if enhanced_recommendations:
                    print("  - 增强建议:")
                    for rec in enhanced_recommendations[:3]:
                        print(f"    • {rec}")
                        
            else:
                print(f"⚠ 增强文件分析失败: {analysis_result.get('error')}")
            
            # 清理
            try:
                os.unlink(test_file)
            except:
                pass
                
        except Exception as e:
            print(f"⚠ 增强文件分析器测试失败: {e}")
        
        print("\n=== 功能特性总结 ===")
        print("✓ 多语言Embedding语义理解")
        print("✓ 智能表格结构检测和分片")
        print("✓ 多模态内容提取（文本+结构+视觉）")
        print("✓ 基于ST-Raptor的层次化分析")
        print("✓ 语义驱动的数据质量检查")
        print("✓ 跨表关系的语义分析")
        print("✓ 智能处理策略推荐")
        
        print("\n=== 使用建议 ===")
        print("1. 安装完整依赖: poetry install")
        print("2. 配置环境变量(.env文件)")
        print("3. 为大文件启用embedding缓存")
        print("4. 使用并行处理提高大表格分析速度")
        
    except ImportError as e:
        print(f"导入错误: {e}")
        print("\n请确保:")
        print("1. 已安装必要依赖: poetry install")
        print("2. 或使用pip: pip install sentence-transformers scikit-learn")
        print("3. Python路径正确配置")
    
    except Exception as e:
        print(f"演示过程中出现错误: {e}")
        import traceback
        traceback.print_exc()


async def create_test_excel():
    """创建测试Excel文件"""
    try:
        import pandas as pd
        
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        # 创建测试数据
        data1 = {
            'ID': range(1, 11),
            '姓名': [f'员工{i}' for i in range(1, 11)],
            '部门': ['销售部', '技术部', '财务部'] * 3 + ['人事部'],
            '薪资': [5000 + i * 500 for i in range(10)]
        }
        
        data2 = {
            '月份': ['1月', '2月', '3月', '4月', '5月'],
            '收入': [10000, 12000, 11000, 13000, 14000],
            '支出': [8000, 9000, 8500, 9500, 10000],
            '利润': [2000, 3000, 2500, 3500, 4000]
        }
        
        # 写入Excel文件
        with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
            pd.DataFrame(data1).to_excel(writer, sheet_name='员工信息', index=False)
            pd.DataFrame(data2).to_excel(writer, sheet_name='财务数据', index=False)
        
        return temp_file.name
        
    except Exception as e:
        print(f"创建测试Excel文件失败: {e}")
        return None


async def create_complex_test_excel():
    """创建复杂的测试Excel文件"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file.close()
        
        wb = Workbook()
        
        # 工作表1: 复杂的销售报表
        ws1 = wb.active
        ws1.title = "销售报表"
        
        # 添加标题
        ws1['A1'] = "2024年度销售业绩报告"
        ws1['A1'].font = Font(size=14, bold=True)
        ws1.merge_cells('A1:E1')
        
        # 添加表头
        headers = ['产品名称', '销售数量', '单价', '总金额', '销售人员']
        for i, header in enumerate(headers, 1):
            ws1.cell(row=2, column=i, value=header)
            ws1.cell(row=2, column=i).font = Font(bold=True)
        
        # 添加数据
        products = ['产品A', '产品B', '产品C', '产品D', '产品E']
        for i, product in enumerate(products, 3):
            ws1.cell(row=i, column=1, value=product)
            ws1.cell(row=i, column=2, value=100 + i * 50)
            ws1.cell(row=i, column=3, value=50 + i * 10)
            ws1.cell(row=i, column=4, value=f"=B{i}*C{i}")  # 公式
            ws1.cell(row=i, column=5, value=f"销售员{i-2}")
        
        # 工作表2: 图表数据
        ws2 = wb.create_sheet("图表数据")
        chart_data = {
            '类别': ['A类', 'B类', 'C类', 'D类'],
            '数值': [25, 35, 20, 20]
        }
        
        for i, (category, value) in enumerate(zip(chart_data['类别'], chart_data['数值']), 1):
            ws2.cell(row=i, column=1, value=category)
            ws2.cell(row=i, column=2, value=value)
        
        # 工作表3: 属性表
        ws3 = wb.create_sheet("系统配置")
        configs = [
            ('系统版本', 'v2.1.0'),
            ('数据库', 'MySQL 8.0'),
            ('服务器', 'Ubuntu 20.04'),
            ('内存', '16GB'),
            ('存储', '1TB SSD')
        ]
        
        for i, (key, value) in enumerate(configs, 1):
            ws3.cell(row=i, column=1, value=key)
            ws3.cell(row=i, column=2, value=value)
        
        wb.save(temp_file.name)
        return temp_file.name
        
    except Exception as e:
        print(f"创建复杂测试Excel文件失败: {e}")
        return None


def main():
    """主函数"""
    print("初始化环境...")
    setup_environment()
    
    # 运行异步演示
    try:
        asyncio.run(demo_enhanced_analysis())
    except KeyboardInterrupt:
        print("\n演示被用户中断")
    except Exception as e:
        print(f"演示运行错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()