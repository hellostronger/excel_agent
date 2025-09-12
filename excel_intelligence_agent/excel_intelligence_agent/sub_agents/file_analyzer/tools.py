"""
Tools for File Analyzer Agent

Specialized tools for Excel file structure analysis and metadata extraction.
"""

import logging
from typing import Dict, Any, List, Tuple, Optional
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xlrd
import re

from google.adk.tools import ToolContext

from ...shared_libraries.utils import setup_logging
from ...shared_libraries.types import FileMetadata
from ...shared_libraries.multimodal_extractor import multimodal_extractor
from ...shared_libraries.tree_partitioner import intelligent_partitioner
from ...shared_libraries.embedding_engine import get_embedding_engine

# Setup logging
logger = setup_logging()


async def analyze_file_structure(
    file_path: str,
    analysis_depth: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """
    Analyze the overall structure of an Excel file using advanced multimodal analysis.
    
    Examines worksheets, data organization, structural patterns, and applies
    ST-Raptor inspired intelligent partitioning.
    """
    try:
        logger.info(f"Starting enhanced file structure analysis: {file_path}")
        
        # Step 1: Basic structure analysis
        path = Path(file_path)
        
        if path.suffix.lower() == '.xlsx':
            basic_structure = await _analyze_xlsx_structure(file_path, analysis_depth)
        elif path.suffix.lower() == '.xls':
            basic_structure = await _analyze_xls_structure(file_path, analysis_depth)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")
        
        # Step 2: Enhanced multimodal content extraction
        try:
            logger.info("Performing multimodal content extraction...")
            multimodal_content = await multimodal_extractor.extract_comprehensive_content(file_path)
            basic_structure["multimodal_analysis"] = multimodal_content
        except Exception as e:
            logger.warning(f"Multimodal analysis failed: {e}")
            basic_structure["multimodal_analysis"] = {"error": str(e), "fallback_used": True}
        
        # Step 3: Intelligent table partitioning
        try:
            logger.info("Performing intelligent table partitioning...")
            partition_analysis = await intelligent_partitioner.analyze_and_partition(file_path)
            basic_structure["partitioning_analysis"] = partition_analysis
            
            # Update structure info with partitioning insights
            if partition_analysis.get("success"):
                basic_structure["intelligent_partitions"] = len(partition_analysis.get("partitions", []))
                basic_structure["processing_strategy"] = partition_analysis.get("processing_strategy", {})
        except Exception as e:
            logger.warning(f"Partitioning analysis failed: {e}")
            basic_structure["partitioning_analysis"] = {"error": str(e), "fallback_used": True}
        
        # Step 4: Semantic analysis if embedding engine is available
        try:
            embedding_engine = get_embedding_engine()
            if embedding_engine:
                logger.info("Performing semantic structure analysis...")
                semantic_insights = await _perform_semantic_structure_analysis(file_path, basic_structure)
                basic_structure["semantic_analysis"] = semantic_insights
            else:
                basic_structure["semantic_analysis"] = {"note": "Embedding engine not available"}
        except Exception as e:
            logger.warning(f"Semantic analysis failed: {e}")
            basic_structure["semantic_analysis"] = {"error": str(e)}
        
        # Step 5: Generate enhanced recommendations
        basic_structure["enhanced_recommendations"] = _generate_enhanced_recommendations(basic_structure)
        
        logger.info("Enhanced file structure analysis completed successfully")
        return basic_structure
            
    except Exception as e:
        logger.error(f"Enhanced file structure analysis failed: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }


async def _analyze_xlsx_structure(file_path: str, analysis_depth: str) -> Dict[str, Any]:
    """Analyze structure of .xlsx files"""
    workbook = load_workbook(file_path, read_only=True, data_only=False)
    
    try:
        structure_info = {
            "success": True,
            "file_format": "xlsx",
            "sheet_count": len(workbook.worksheets),
            "sheet_names": [ws.title for ws in workbook.worksheets],
            "sheets_analysis": [],
            "total_data_cells": 0,
            "has_formulas": False,
            "has_charts": False,
            "has_pivot_tables": False
        }
        
        for worksheet in workbook.worksheets:
            sheet_analysis = await _analyze_worksheet_structure(worksheet, analysis_depth)
            structure_info["sheets_analysis"].append(sheet_analysis)
            structure_info["total_data_cells"] += sheet_analysis.get("data_cell_count", 0)
            
            # Check for advanced features
            if sheet_analysis.get("has_formulas", False):
                structure_info["has_formulas"] = True
            if sheet_analysis.get("has_charts", False):
                structure_info["has_charts"] = True
        
        # Analyze cross-sheet relationships
        structure_info["cross_sheet_references"] = await _find_cross_sheet_references(workbook)
        
        return structure_info
        
    finally:
        workbook.close()


async def _analyze_xls_structure(file_path: str, analysis_depth: str) -> Dict[str, Any]:
    """Analyze structure of .xls files"""
    workbook = xlrd.open_workbook(file_path)
    
    structure_info = {
        "success": True,
        "file_format": "xls",
        "sheet_count": len(workbook.sheet_names()),
        "sheet_names": workbook.sheet_names(),
        "sheets_analysis": [],
        "total_data_cells": 0,
        "has_formulas": False,
        "has_charts": False,
        "has_pivot_tables": False
    }
    
    for sheet_name in workbook.sheet_names():
        worksheet = workbook.sheet_by_name(sheet_name)
        sheet_analysis = await _analyze_xls_worksheet_structure(worksheet, analysis_depth)
        structure_info["sheets_analysis"].append(sheet_analysis)
        structure_info["total_data_cells"] += sheet_analysis.get("data_cell_count", 0)
        
        if sheet_analysis.get("has_formulas", False):
            structure_info["has_formulas"] = True
    
    return structure_info


async def _analyze_worksheet_structure(worksheet, analysis_depth: str) -> Dict[str, Any]:
    """Analyze structure of an individual worksheet (.xlsx)"""
    
    # Get worksheet dimensions
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    
    # Count actual data cells
    data_cell_count = 0
    formula_count = 0
    
    # Sample cells for analysis (don't read entire large sheets)
    sample_size = min(max_row, 1000) if analysis_depth == "comprehensive" else min(max_row, 100)
    
    for row in worksheet.iter_rows(max_row=sample_size, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                data_cell_count += 1
                if str(cell.value).startswith('='):
                    formula_count += 1
    
    # Estimate total data cells if we sampled
    if sample_size < max_row:
        scaling_factor = max_row / sample_size
        data_cell_count = int(data_cell_count * scaling_factor)
        formula_count = int(formula_count * scaling_factor)
    
    # Identify data patterns
    data_patterns = await _identify_data_patterns(worksheet, sample_size)
    
    return {
        "sheet_name": worksheet.title,
        "max_row": max_row,
        "max_column": max_col,
        "data_cell_count": data_cell_count,
        "formula_count": formula_count,
        "has_formulas": formula_count > 0,
        "has_charts": len(worksheet._charts) > 0 if hasattr(worksheet, '_charts') else False,
        "data_patterns": data_patterns,
        "estimated_data_density": data_cell_count / (max_row * max_col) if (max_row * max_col) > 0 else 0
    }


async def _analyze_xls_worksheet_structure(worksheet, analysis_depth: str) -> Dict[str, Any]:
    """Analyze structure of an individual worksheet (.xls)"""
    
    max_row = worksheet.nrows
    max_col = worksheet.ncols
    
    data_cell_count = 0
    formula_count = 0
    
    # Sample for analysis
    sample_size = min(max_row, 1000) if analysis_depth == "comprehensive" else min(max_row, 100)
    
    for row_idx in range(sample_size):
        for col_idx in range(max_col):
            try:
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value is not None and cell.value != '':
                    data_cell_count += 1
                    # XLS formula detection
                    if cell.ctype == xlrd.XL_CELL_FORMULA:
                        formula_count += 1
            except:
                continue
    
    # Scale estimates if sampled
    if sample_size < max_row:
        scaling_factor = max_row / sample_size
        data_cell_count = int(data_cell_count * scaling_factor)
        formula_count = int(formula_count * scaling_factor)
    
    return {
        "sheet_name": worksheet.name,
        "max_row": max_row,
        "max_column": max_col,
        "data_cell_count": data_cell_count,
        "formula_count": formula_count,
        "has_formulas": formula_count > 0,
        "has_charts": False,  # Charts detection limited in XLS
        "estimated_data_density": data_cell_count / (max_row * max_col) if (max_row * max_col) > 0 else 0
    }


async def _identify_data_patterns(worksheet, sample_size: int) -> Dict[str, Any]:
    """Identify data organization patterns in a worksheet"""
    patterns = {
        "has_headers": False,
        "header_row": None,
        "data_tables": [],
        "calculation_areas": [],
        "empty_regions": []
    }
    
    # Check first few rows for headers
    for row_num in range(1, min(6, sample_size)):
        row_values = []
        row = list(worksheet.iter_rows(min_row=row_num, max_row=row_num, max_col=min(worksheet.max_column, 20)))[0]
        
        for cell in row:
            if cell.value is not None:
                row_values.append(str(cell.value))
        
        # Header detection heuristics
        if len(row_values) > 1 and _looks_like_headers(row_values):
            patterns["has_headers"] = True
            patterns["header_row"] = row_num
            break
    
    return patterns


def _looks_like_headers(row_values: List[str]) -> bool:
    """Heuristic to determine if a row contains headers"""
    if not row_values:
        return False
    
    # Check for header characteristics
    text_count = sum(1 for val in row_values if not _is_numeric(val))
    total_count = len(row_values)
    
    # Most values should be text for headers
    if text_count / total_count > 0.7:
        return True
    
    # Check for common header patterns
    header_keywords = ['id', 'name', 'date', 'amount', 'total', 'count', 'value', 'description']
    keyword_matches = sum(1 for val in row_values if any(keyword in val.lower() for keyword in header_keywords))
    
    return keyword_matches > 0


def _is_numeric(value: str) -> bool:
    """Check if a string represents a numeric value"""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


async def _find_cross_sheet_references(workbook) -> List[Dict[str, Any]]:
    """Find references between worksheets"""
    references = []
    
    for worksheet in workbook.worksheets:
        # Look for formulas that reference other sheets
        for row in worksheet.iter_rows(max_row=min(worksheet.max_row, 100)):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # Simple regex to find sheet references like 'Sheet1!'
                    sheet_refs = re.findall(r"'?([^'!]+)'?!", cell.value)
                    for ref_sheet in sheet_refs:
                        if ref_sheet != worksheet.title:
                            references.append({
                                "source_sheet": worksheet.title,
                                "target_sheet": ref_sheet,
                                "cell_reference": f"{get_column_letter(cell.column)}{cell.row}",
                                "formula": cell.value[:100]  # Truncate long formulas
                            })
    
    return references


async def extract_sheet_metadata(
    file_path: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """
    Extract detailed metadata for each worksheet.
    
    Provides comprehensive information about each sheet's characteristics.
    """
    try:
        path = Path(file_path)
        
        if path.suffix.lower() == '.xlsx':
            return await _extract_xlsx_sheet_metadata(file_path)
        elif path.suffix.lower() == '.xls':
            return await _extract_xls_sheet_metadata(file_path)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")
            
    except Exception as e:
        logger.error(f"Sheet metadata extraction failed: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }


async def _extract_xlsx_sheet_metadata(file_path: str) -> Dict[str, Any]:
    """Extract metadata from .xlsx worksheets"""
    workbook = load_workbook(file_path, read_only=True)
    
    try:
        metadata = {
            "success": True,
            "sheets": [],
            "data_quality_indicators": {
                "has_quality_issues": False,
                "issues_found": []
            }
        }
        
        for worksheet in workbook.worksheets:
            sheet_meta = {
                "name": worksheet.title,
                "dimensions": {
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column
                },
                "estimated_purpose": await _infer_sheet_purpose(worksheet),
                "data_characteristics": await _analyze_data_characteristics(worksheet)
            }
            
            # Check for quality issues
            if sheet_meta["data_characteristics"].get("sparse_data", False):
                metadata["data_quality_indicators"]["has_quality_issues"] = True
                metadata["data_quality_indicators"]["issues_found"].append(
                    f"Sparse data detected in sheet '{worksheet.title}'"
                )
            
            metadata["sheets"].append(sheet_meta)
        
        return metadata
        
    finally:
        workbook.close()


async def _extract_xls_sheet_metadata(file_path: str) -> Dict[str, Any]:
    """Extract metadata from .xls worksheets"""
    workbook = xlrd.open_workbook(file_path)
    
    metadata = {
        "success": True,
        "sheets": [],
        "data_quality_indicators": {
            "has_quality_issues": False,
            "issues_found": []
        }
    }
    
    for sheet_name in workbook.sheet_names():
        worksheet = workbook.sheet_by_name(sheet_name)
        
        sheet_meta = {
            "name": sheet_name,
            "dimensions": {
                "max_row": worksheet.nrows,
                "max_column": worksheet.ncols
            },
            "estimated_purpose": await _infer_xls_sheet_purpose(worksheet),
            "data_characteristics": await _analyze_xls_data_characteristics(worksheet)
        }
        
        metadata["sheets"].append(sheet_meta)
    
    return metadata


async def _infer_sheet_purpose(worksheet) -> str:
    """Infer the likely purpose of a worksheet based on its structure and content"""
    
    # Sample first few rows to analyze content
    sample_data = []
    for row in worksheet.iter_rows(max_row=min(10, worksheet.max_row), values_only=True):
        sample_data.append([str(cell) if cell is not None else '' for cell in row])
    
    if not sample_data:
        return "empty"
    
    # Analysis heuristics
    first_row = sample_data[0] if sample_data else []
    
    # Look for keywords that indicate purpose
    content_text = ' '.join([' '.join(row) for row in sample_data]).lower()
    
    if any(keyword in content_text for keyword in ['summary', 'total', 'overview']):
        return "summary"
    elif any(keyword in content_text for keyword in ['data', 'records', 'entries']):
        return "data_table"
    elif any(keyword in content_text for keyword in ['calculation', 'formula', 'compute']):
        return "calculations"
    elif len(first_row) > 5 and most_are_text(first_row):
        return "data_table"
    else:
        return "general"


async def _infer_xls_sheet_purpose(worksheet) -> str:
    """Infer purpose for XLS worksheets"""
    # Simplified version for XLS
    if worksheet.nrows < 2:
        return "empty"
    elif worksheet.ncols > 5:
        return "data_table"
    else:
        return "general"


def most_are_text(row_values: List[str]) -> bool:
    """Check if most values in a row are text (not numeric)"""
    if not row_values:
        return False
    
    text_count = sum(1 for val in row_values if val and not _is_numeric(val))
    return text_count / len(row_values) > 0.6


async def _analyze_data_characteristics(worksheet) -> Dict[str, Any]:
    """Analyze data characteristics of a worksheet"""
    
    total_cells = worksheet.max_row * worksheet.max_column
    
    # Count non-empty cells in sample
    sample_size = min(worksheet.max_row, 200)
    non_empty_count = 0
    
    for row in worksheet.iter_rows(max_row=sample_size):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != '':
                non_empty_count += 1
    
    # Estimate density
    if sample_size < worksheet.max_row:
        scaling_factor = worksheet.max_row / sample_size
        estimated_non_empty = int(non_empty_count * scaling_factor)
    else:
        estimated_non_empty = non_empty_count
    
    density = estimated_non_empty / total_cells if total_cells > 0 else 0
    
    return {
        "data_density": density,
        "sparse_data": density < 0.1,
        "estimated_non_empty_cells": estimated_non_empty,
        "likely_tabular": density > 0.3 and worksheet.max_column > 2
    }


async def _analyze_xls_data_characteristics(worksheet) -> Dict[str, Any]:
    """Analyze data characteristics for XLS worksheets"""
    
    total_cells = worksheet.nrows * worksheet.ncols
    non_empty_count = 0
    
    # Sample analysis
    sample_size = min(worksheet.nrows, 200)
    
    for row_idx in range(sample_size):
        for col_idx in range(worksheet.ncols):
            try:
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value is not None and str(cell.value).strip() != '':
                    non_empty_count += 1
            except:
                continue
    
    # Estimate density
    if sample_size < worksheet.nrows:
        scaling_factor = worksheet.nrows / sample_size
        estimated_non_empty = int(non_empty_count * scaling_factor)
    else:
        estimated_non_empty = non_empty_count
    
    density = estimated_non_empty / total_cells if total_cells > 0 else 0
    
    return {
        "data_density": density,
        "sparse_data": density < 0.1,
        "estimated_non_empty_cells": estimated_non_empty,
        "likely_tabular": density > 0.3 and worksheet.ncols > 2
    }


async def assess_file_complexity(
    file_path: str,
    tool_context: ToolContext
) -> Dict[str, Any]:
    """
    Assess the complexity of an Excel file for processing planning.
    
    Determines processing requirements and resource needs.
    """
    try:
        # Get basic file information
        path = Path(file_path)
        file_size = path.stat().st_size
        
        # Get structure information from context if available
        structure_info = tool_context.state.get("structural_analysis", {})
        
        complexity_factors = {
            "file_size_mb": file_size / (1024 * 1024),
            "sheet_count": structure_info.get("sheet_count", 0),
            "total_data_cells": structure_info.get("total_data_cells", 0),
            "has_formulas": structure_info.get("has_formulas", False),
            "has_cross_references": len(structure_info.get("cross_sheet_references", [])) > 0,
            "has_charts": structure_info.get("has_charts", False)
        }
        
        # Calculate complexity score
        complexity_score = _calculate_complexity_score(complexity_factors)
        
        # Determine complexity level
        if complexity_score > 7:
            complexity_level = "high"
            processing_time_estimate = "5-10 minutes"
            memory_requirement = "high"
        elif complexity_score > 4:
            complexity_level = "medium" 
            processing_time_estimate = "2-5 minutes"
            memory_requirement = "medium"
        else:
            complexity_level = "low"
            processing_time_estimate = "1-2 minutes"
            memory_requirement = "low"
        
        return {
            "success": True,
            "complexity_level": complexity_level,
            "complexity_score": complexity_score,
            "complexity_factors": complexity_factors,
            "processing_estimate": {
                "time_estimate": processing_time_estimate,
                "memory_requirement": memory_requirement,
                "recommended_parallel_agents": min(complexity_score, 4)
            },
            "confidence": 0.8
        }
        
    except Exception as e:
        logger.error(f"Complexity assessment failed: {str(e)}")
        return {
            "success": False,
            "error": str(e),
            "complexity_level": "unknown",
            "confidence": 0.0
        }


def _calculate_complexity_score(factors: Dict[str, Any]) -> float:
    """Calculate numerical complexity score based on various factors"""
    score = 0.0
    
    # File size contribution
    size_mb = factors.get("file_size_mb", 0)
    if size_mb > 10:
        score += 3
    elif size_mb > 5:
        score += 2
    elif size_mb > 1:
        score += 1
    
    # Sheet count contribution
    sheet_count = factors.get("sheet_count", 0)
    if sheet_count > 10:
        score += 2
    elif sheet_count > 5:
        score += 1
    
    # Data cells contribution
    data_cells = factors.get("total_data_cells", 0)
    if data_cells > 100000:
        score += 3
    elif data_cells > 10000:
        score += 2
    elif data_cells > 1000:
        score += 1
    
    # Advanced features
    if factors.get("has_formulas", False):
        score += 1
    if factors.get("has_cross_references", False):
        score += 1
    if factors.get("has_charts", False):
        score += 0.5
    
    # Enhanced factors from ST-Raptor analysis
    if factors.get("intelligent_partitions", 0) > 5:
        score += 1
    if factors.get("has_multimodal_content", False):
        score += 1
    if factors.get("semantic_complexity", 0) > 0.7:
        score += 1
    
    return min(score, 10.0)  # Cap at 10


async def _perform_semantic_structure_analysis(file_path: str, structure_info: Dict[str, Any]) -> Dict[str, Any]:
    """
    使用embedding模型进行语义结构分析
    """
    try:
        embedding_engine = get_embedding_engine()
        if not embedding_engine:
            return {"error": "Embedding engine not available"}
        
        semantic_insights = {
            "sheet_semantic_relationships": [],
            "cross_sheet_semantic_similarity": {},
            "content_categorization": {},
            "semantic_coherence_score": 0.0
        }
        
        # 从multimodal分析中提取文本内容
        multimodal_data = structure_info.get("multimodal_analysis", {})
        if multimodal_data.get("success"):
            sheets_data = multimodal_data.get("sheets", [])
            
            # 分析工作表之间的语义关系
            sheet_names = [sheet.get("name", "") for sheet in sheets_data]
            if len(sheet_names) > 1:
                # 计算工作表名称的语义相似度
                similarity_results = embedding_engine.batch_similarity_search(
                    sheet_names, sheet_names, top_k=min(3, len(sheet_names)-1)
                )
                
                for i, (sheet_name, similar_sheets) in enumerate(zip(sheet_names, similarity_results)):
                    # 过滤掉自己
                    filtered_similar = [(name, score) for name, score in similar_sheets if name != sheet_name]
                    if filtered_similar:
                        semantic_insights["sheet_semantic_relationships"].append({
                            "sheet": sheet_name,
                            "similar_sheets": filtered_similar[:2]  # Top 2
                        })
            
            # 分析内容分类
            all_content_samples = []
            sheet_content_mapping = {}
            
            for sheet in sheets_data:
                sheet_name = sheet.get("name", "")
                content_analysis = sheet.get("content_analysis", {})
                cell_samples = content_analysis.get("cell_samples", [])
                
                sheet_content = [sample.get("value", "") for sample in cell_samples if sample.get("value")]
                if sheet_content:
                    all_content_samples.extend(sheet_content)
                    sheet_content_mapping[sheet_name] = sheet_content
            
            # 使用embedding进行内容分类
            if all_content_samples:
                categories = ["财务数据", "人员信息", "时间数据", "产品信息", "统计数据", "文本描述"]
                
                try:
                    categorized_content = embedding_engine.categorize_column_content(
                        all_content_samples[:50],  # 限制样本数量
                        categories
                    )
                    semantic_insights["content_categorization"] = categorized_content
                except Exception as e:
                    logger.warning(f"Content categorization failed: {e}")
        
        # 计算整体语义一致性
        partitioning_data = structure_info.get("partitioning_analysis", {})
        if partitioning_data.get("success"):
            sheets = partitioning_data.get("sheets", [])
            complexity_scores = [sheet.get("complexity_level", "medium") for sheet in sheets]
            
            # 简单的一致性评分
            unique_complexities = len(set(complexity_scores))
            total_sheets = len(sheets)
            
            if total_sheets > 0:
                consistency_score = 1.0 - (unique_complexities - 1) / max(total_sheets - 1, 1)
                semantic_insights["semantic_coherence_score"] = consistency_score
        
        return semantic_insights
        
    except Exception as e:
        logger.error(f"Semantic structure analysis failed: {e}")
        return {"error": str(e)}


def _generate_enhanced_recommendations(structure_info: Dict[str, Any]) -> List[str]:
    """
    基于增强分析生成建议
    """
    recommendations = []
    
    # 基于分片分析的建议
    partitioning_data = structure_info.get("partitioning_analysis", {})
    if partitioning_data.get("success"):
        strategy = partitioning_data.get("processing_strategy", {})
        overall_strategy = strategy.get("overall_strategy", "")
        
        if overall_strategy == "hierarchical_multimodal":
            recommendations.append("建议使用层次化多模态处理策略，启用并行Agent处理")
        elif overall_strategy == "parallel_processing":
            recommendations.append("建议使用并行处理策略，可同时分析多个表格分片")
        
        total_partitions = strategy.get("total_partitions", 0)
        if total_partitions > 10:
            recommendations.append(f"检测到{total_partitions}个分片，建议启用智能分片缓存机制")
    
    # 基于多模态分析的建议
    multimodal_data = structure_info.get("multimodal_analysis", {})
    if multimodal_data.get("success"):
        global_insights = multimodal_data.get("global_insights", {})
        
        if global_insights.get("has_charts"):
            recommendations.append("检测到图表元素，建议启用视觉内容分析")
        if global_insights.get("has_images"):
            recommendations.append("检测到图像内容，建议使用多模态处理管道")
        if global_insights.get("has_formulas"):
            recommendations.append("检测到公式，建议深度分析计算依赖关系")
    
    # 基于语义分析的建议
    semantic_data = structure_info.get("semantic_analysis", {})
    if not semantic_data.get("error"):
        coherence_score = semantic_data.get("semantic_coherence_score", 0.0)
        
        if coherence_score < 0.3:
            recommendations.append("语义一致性较低，建议检查数据质量和结构规范性")
        elif coherence_score > 0.8:
            recommendations.append("语义一致性良好，适合使用语义驱动的智能分析")
        
        relationships = semantic_data.get("sheet_semantic_relationships", [])
        if len(relationships) > 0:
            recommendations.append("发现工作表间语义关联，建议启用跨表关系分析")
    
    # 性能优化建议
    if structure_info.get("intelligent_partitions", 0) > 8:
        recommendations.append("建议使用embedding缓存加速重复分析")
    
    # 默认建议
    if not recommendations:
        recommendations.append("使用标准多Agent协作分析流程")
        recommendations.append("建议启用embedding辅助的语义理解")
    
    return recommendations