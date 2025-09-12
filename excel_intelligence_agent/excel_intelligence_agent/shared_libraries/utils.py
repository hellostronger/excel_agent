"""
Shared utility functions for Excel Intelligence Agent System
"""

import os
import logging
import asyncio
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
import xlrd

from .types import FileMetadata, AnalysisDepth, ColumnProfile, DataQualityLevel
from .constants import (
    MAX_FILE_SIZE_MB, 
    EXCEL_DATA_TYPES,
    EXCELLENT_QUALITY_THRESHOLD,
    GOOD_QUALITY_THRESHOLD, 
    FAIR_QUALITY_THRESHOLD,
    POOR_QUALITY_THRESHOLD
)


def setup_logging(level: str = "INFO") -> logging.Logger:
    """Setup logging configuration for the system"""
    logging.basicConfig(
        level=getattr(logging, level.upper()),
        format='%(asctime)s | %(name)s | %(levelname)s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    return logging.getLogger("excel_intelligence_agent")


def load_environment_variables() -> Dict[str, Any]:
    """Load and validate environment variables"""
    
    # 模型提供商选择
    model_provider = os.getenv("MODEL_PROVIDER", "siliconflow").lower()
    
    # 根据提供商确定必需的变量
    if model_provider == "google":
        required_vars = [
            "GOOGLE_CLOUD_PROJECT",
            "GOOGLE_CLOUD_LOCATION", 
            "ROOT_AGENT_MODEL"
        ]
    elif model_provider == "siliconflow":
        required_vars = [
            "SILICONFLOW_API_KEY",
            "ROOT_AGENT_MODEL"
        ]
    else:
        raise ValueError(f"Unsupported model provider: {model_provider}. Use 'google' or 'siliconflow'")
    
    env_config = {}
    missing_vars = []
    
    for var in required_vars:
        value = os.getenv(var)
        if value is None:
            missing_vars.append(var)
        else:
            env_config[var] = value
    
    if missing_vars:
        raise ValueError(f"Missing required environment variables for {model_provider} provider: {missing_vars}")
    
    # 通用配置
    env_config.update({
        "MODEL_PROVIDER": model_provider,
        "MAX_PARALLEL_AGENTS": int(os.getenv("MAX_PARALLEL_AGENTS", "4")),
        "ENABLE_CONCURRENT_ANALYSIS": os.getenv("ENABLE_CONCURRENT_ANALYSIS", "true").lower() == "true",
        "FILE_PREPARATION_TIMEOUT": int(os.getenv("FILE_PREPARATION_TIMEOUT", "120")),
        "PARALLEL_ANALYSIS_TIMEOUT": int(os.getenv("PARALLEL_ANALYSIS_TIMEOUT", "300")),
        "RESPONSE_GENERATION_TIMEOUT": int(os.getenv("RESPONSE_GENERATION_TIMEOUT", "60"))
    })
    
    # 模型配置 - 根据提供商使用不同的默认值
    if model_provider == "siliconflow":
        env_config.update({
            "ORCHESTRATOR_MODEL": os.getenv("ORCHESTRATOR_MODEL", "qwen-max"),
            "WORKER_MODEL": os.getenv("WORKER_MODEL", "qwen-turbo"),
            "SILICONFLOW_BASE_URL": os.getenv("SILICONFLOW_BASE_URL", "https://api.siliconflow.cn/v1"),
            "SILICONFLOW_TIMEOUT": int(os.getenv("SILICONFLOW_TIMEOUT", "300"))
        })
    else:  # google
        env_config.update({
            "ORCHESTRATOR_MODEL": os.getenv("ORCHESTRATOR_MODEL", "gemini-2.5-pro"),
            "WORKER_MODEL": os.getenv("WORKER_MODEL", "gemini-2.5-flash"),
        })
    
    return env_config


def create_model_client(provider: str = None):
    """创建模型客户端"""
    if provider is None:
        provider = os.getenv("MODEL_PROVIDER", "siliconflow").lower()
    
    if provider == "siliconflow":
        from .siliconflow_client import create_siliconflow_client
        return create_siliconflow_client()
    elif provider == "google":
        # Google ADK客户端已经内置
        return None
    else:
        raise ValueError(f"Unsupported model provider: {provider}")


async def validate_excel_file(file_path: str) -> Tuple[bool, str]:
    """Validate Excel file accessibility and format"""
    path = Path(file_path)
    
    # Check file existence
    if not path.exists():
        return False, f"File not found: {file_path}"
    
    # Check file size
    file_size_mb = path.stat().st_size / (1024 * 1024)
    if file_size_mb > MAX_FILE_SIZE_MB:
        return False, f"File too large: {file_size_mb:.1f}MB (max: {MAX_FILE_SIZE_MB}MB)"
    
    # Check file format
    valid_extensions = ['.xlsx', '.xls', '.xlsm']
    if path.suffix.lower() not in valid_extensions:
        return False, f"Invalid file format: {path.suffix} (supported: {valid_extensions})"
    
    # Try to read the file
    try:
        if path.suffix.lower() == '.xlsx':
            workbook = load_workbook(str(path), read_only=True)
            workbook.close()
        elif path.suffix.lower() == '.xls':
            xlrd.open_workbook(str(path))
        return True, "File validation successful"
    except Exception as e:
        return False, f"File format error: {str(e)}"


async def extract_basic_metadata(file_path: str) -> FileMetadata:
    """Extract basic metadata from Excel file"""
    path = Path(file_path)
    stat = path.stat()
    
    # Count sheets
    sheets_count = 0
    try:
        if path.suffix.lower() == '.xlsx':
            workbook = load_workbook(str(path), read_only=True)
            sheets_count = len(workbook.worksheets)
            workbook.close()
        elif path.suffix.lower() == '.xls':
            workbook = xlrd.open_workbook(str(path))
            sheets_count = len(workbook.sheet_names())
    except Exception:
        sheets_count = 0
    
    return FileMetadata(
        file_path=str(path),
        file_size=stat.st_size,
        sheets_count=sheets_count,
        creation_time=datetime.fromtimestamp(stat.st_ctime),
        modification_time=datetime.fromtimestamp(stat.st_mtime),
        extraction_timestamp=datetime.now(),
        analysis_depth=AnalysisDepth.BASIC
    )


def infer_data_type(series: pd.Series) -> Tuple[str, float]:
    """Infer data type of a pandas Series with confidence score"""
    if series.empty:
        return "blank", 1.0
    
    # Remove null values for type inference
    non_null_series = series.dropna()
    if non_null_series.empty:
        return "blank", 1.0
    
    total_count = len(non_null_series)
    
    # Check for boolean
    if series.dtype == 'bool' or all(val in [True, False, 0, 1, "true", "false", "TRUE", "FALSE"] 
                                     for val in non_null_series):
        return "boolean", 0.95
    
    # Check for numeric types
    if pd.api.types.is_numeric_dtype(series):
        if pd.api.types.is_integer_dtype(series):
            return "integer", 0.9
        else:
            return "float", 0.9
    
    # Convert to string for pattern analysis
    str_series = non_null_series.astype(str)
    
    # Check for dates
    date_patterns = 0
    for val in str_series.head(min(100, len(str_series))):
        try:
            pd.to_datetime(val)
            date_patterns += 1
        except:
            pass
    
    if date_patterns / min(100, total_count) > 0.7:
        return "date", 0.8
    
    # Check for currency
    currency_patterns = sum(1 for val in str_series.head(100) 
                           if any(symbol in str(val) for symbol in ['$', '€', '£', '¥']))
    if currency_patterns / min(100, total_count) > 0.5:
        return "currency", 0.7
    
    # Check for percentage
    percentage_patterns = sum(1 for val in str_series.head(100) if '%' in str(val))
    if percentage_patterns / min(100, total_count) > 0.5:
        return "percentage", 0.7
    
    # Default to text
    return "text", 0.6


def calculate_quality_score(column: pd.Series) -> Tuple[float, DataQualityLevel]:
    """Calculate quality score and level for a data column"""
    if column.empty:
        return 0.0, DataQualityLevel.CRITICAL
    
    total_count = len(column)
    null_count = column.isnull().sum()
    non_null_count = total_count - null_count
    
    if non_null_count == 0:
        return 0.0, DataQualityLevel.CRITICAL
    
    # Base score from completeness
    completeness_score = non_null_count / total_count
    
    # Consistency score (based on data type uniformity)
    non_null_series = column.dropna()
    data_type, type_confidence = infer_data_type(non_null_series)
    consistency_score = type_confidence
    
    # Uniqueness penalty for excessive duplicates (except for categorical data)
    unique_count = column.nunique()
    if unique_count < non_null_count * 0.1 and data_type not in ["boolean", "category"]:
        uniqueness_penalty = 0.2
    else:
        uniqueness_penalty = 0.0
    
    # Calculate overall quality score
    quality_score = (completeness_score * 0.4 + 
                    consistency_score * 0.4 + 
                    (1 - uniqueness_penalty) * 0.2)
    
    # Determine quality level
    if quality_score >= EXCELLENT_QUALITY_THRESHOLD:
        quality_level = DataQualityLevel.EXCELLENT
    elif quality_score >= GOOD_QUALITY_THRESHOLD:
        quality_level = DataQualityLevel.GOOD
    elif quality_score >= FAIR_QUALITY_THRESHOLD:
        quality_level = DataQualityLevel.FAIR
    elif quality_score >= POOR_QUALITY_THRESHOLD:
        quality_level = DataQualityLevel.POOR
    else:
        quality_level = DataQualityLevel.CRITICAL
    
    return quality_score, quality_level


def detect_column_anomalies(column: pd.Series, data_type: str) -> List[str]:
    """Detect anomalies in a data column"""
    anomalies = []
    
    if column.empty:
        anomalies.append("Column is empty")
        return anomalies
    
    # Check for high null percentage
    null_percentage = column.isnull().sum() / len(column)
    if null_percentage > 0.5:
        anomalies.append(f"High null percentage: {null_percentage:.1%}")
    
    # Check for excessive duplicates
    unique_count = column.nunique()
    if unique_count < len(column) * 0.05 and data_type not in ["boolean"]:
        anomalies.append(f"Very low uniqueness: {unique_count} unique values in {len(column)} rows")
    
    # Type-specific anomalies
    if data_type in ["integer", "float"]:
        # Check for outliers using IQR method
        numeric_column = pd.to_numeric(column, errors='coerce').dropna()
        if len(numeric_column) > 0:
            Q1 = numeric_column.quantile(0.25)
            Q3 = numeric_column.quantile(0.75)
            IQR = Q3 - Q1
            outliers = numeric_column[(numeric_column < Q1 - 1.5 * IQR) | 
                                    (numeric_column > Q3 + 1.5 * IQR)]
            if len(outliers) > 0:
                anomalies.append(f"Potential outliers detected: {len(outliers)} values")
    
    if data_type == "text":
        # Check for inconsistent formatting
        non_null_values = column.dropna().astype(str)
        if len(non_null_values) > 0:
            lengths = non_null_values.str.len()
            if lengths.std() > lengths.mean():
                anomalies.append("Inconsistent text length patterns")
    
    return anomalies


def sample_column_values(column: pd.Series, max_samples: int = 10) -> List[Any]:
    """Get representative sample values from a column"""
    if column.empty:
        return []
    
    # Remove nulls for sampling
    non_null_column = column.dropna()
    if non_null_column.empty:
        return []
    
    # Get unique values first, then sample if needed
    unique_values = non_null_column.unique()
    
    if len(unique_values) <= max_samples:
        return unique_values.tolist()
    else:
        # Sample from most frequent values
        value_counts = non_null_column.value_counts()
        return value_counts.head(max_samples).index.tolist()


async def run_with_timeout(coro, timeout_seconds: int):
    """Run an async function with timeout"""
    try:
        return await asyncio.wait_for(coro, timeout=timeout_seconds)
    except asyncio.TimeoutError:
        raise TimeoutError(f"Operation timed out after {timeout_seconds} seconds")


def format_file_size(size_bytes: int) -> str:
    """Format file size in human readable format"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.1f} TB"


def truncate_text(text: str, max_length: int = 100) -> str:
    """Truncate text to specified length with ellipsis"""
    if len(text) <= max_length:
        return text
    return text[:max_length-3] + "..."