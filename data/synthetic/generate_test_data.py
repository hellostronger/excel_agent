"""Generate synthetic Excel test data for validation."""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
import random

def generate_sales_data():
    """Generate synthetic sales data."""
    np.random.seed(42)
    
    # Generate date range
    start_date = datetime(2024, 1, 1)
    end_date = datetime(2024, 12, 31)
    date_range = pd.date_range(start_date, end_date, freq='D')
    
    # Product data
    products = [
        'Laptop', 'Desktop', 'Tablet', 'Phone', 'Monitor', 
        'Keyboard', 'Mouse', 'Headphones', 'Speaker', 'Camera'
    ]
    
    categories = ['Electronics', 'Accessories', 'Computing']
    regions = ['North', 'South', 'East', 'West', 'Central']
    
    # Generate sales data
    n_records = 1000
    data = []
    
    for _ in range(n_records):
        date = random.choice(date_range)
        product = random.choice(products)
        category = random.choice(categories)
        region = random.choice(regions)
        quantity = random.randint(1, 50)
        unit_price = round(random.uniform(50, 2000), 2)
        total_sales = round(quantity * unit_price, 2)
        
        # Add some seasonal trends
        month = date.month
        seasonal_multiplier = 1 + 0.3 * np.sin(2 * np.pi * month / 12)
        total_sales *= seasonal_multiplier
        total_sales = round(total_sales, 2)
        
        data.append({
            'Date': date,
            'Product': product,
            'Category': category,
            'Region': region,
            'Quantity': quantity,
            'Unit_Price': unit_price,
            'Total_Sales': total_sales,
            'Sales_Rep': f"Rep_{random.randint(1, 20):02d}",
            'Customer_Type': random.choice(['B2B', 'B2C'])
        })
    
    return pd.DataFrame(data)


def generate_customer_data():
    """Generate synthetic customer data."""
    np.random.seed(42)
    
    n_customers = 200
    data = []
    
    for i in range(n_customers):
        customer_id = f"CUST_{i+1:04d}"
        company_name = f"Company {chr(65 + i % 26)}{i+1}"
        contact_name = f"Contact {random.choice(['John', 'Jane', 'Bob', 'Alice', 'Charlie'])} {random.choice(['Smith', 'Johnson', 'Brown', 'Davis', 'Wilson'])}"
        email = f"{contact_name.lower().replace(' ', '.')}@{company_name.lower().replace(' ', '')}.com"
        
        data.append({
            'Customer_ID': customer_id,
            'Company_Name': company_name,
            'Contact_Name': contact_name,
            'Email': email,
            'Phone': f"+1-{random.randint(200, 999)}-{random.randint(100, 999)}-{random.randint(1000, 9999)}",
            'Region': random.choice(['North', 'South', 'East', 'West', 'Central']),
            'Industry': random.choice(['Technology', 'Healthcare', 'Finance', 'Education', 'Retail']),
            'Company_Size': random.choice(['Small', 'Medium', 'Large']),
            'Annual_Revenue': random.randint(100000, 50000000),
            'Registration_Date': datetime(2023, 1, 1) + timedelta(days=random.randint(0, 730))
        })
    
    return pd.DataFrame(data)


def generate_inventory_data():
    """Generate synthetic inventory data."""
    products = [
        'Laptop', 'Desktop', 'Tablet', 'Phone', 'Monitor', 
        'Keyboard', 'Mouse', 'Headphones', 'Speaker', 'Camera'
    ]
    
    data = []
    for product in products:
        data.append({
            'Product_ID': f"PROD_{products.index(product)+1:03d}",
            'Product_Name': product,
            'Category': 'Electronics' if product in ['Laptop', 'Desktop', 'Tablet', 'Phone'] else 'Accessories',
            'Stock_Quantity': random.randint(10, 500),
            'Reorder_Level': random.randint(20, 100),
            'Unit_Cost': round(random.uniform(30, 1500), 2),
            'Supplier': f"Supplier_{random.randint(1, 5)}",
            'Last_Updated': datetime.now() - timedelta(days=random.randint(0, 30))
        })
    
    return pd.DataFrame(data)


def create_test_excel_files():
    """Create test Excel files with multiple scenarios."""
    output_dir = Path(__file__).parent
    output_dir.mkdir(exist_ok=True)
    
    # Scenario 1: Single table analysis
    print("Creating single table test file...")
    sales_data = generate_sales_data()
    
    with pd.ExcelWriter(output_dir / "single_table_sales.xlsx", engine='openpyxl') as writer:
        sales_data.to_excel(writer, sheet_name='Sales_Data', index=False)
        
        # Add some summary data
        monthly_summary = sales_data.groupby(sales_data['Date'].dt.to_period('M')).agg({
            'Total_Sales': 'sum',
            'Quantity': 'sum'
        }).reset_index()
        monthly_summary['Date'] = monthly_summary['Date'].astype(str)
        monthly_summary.to_excel(writer, sheet_name='Monthly_Summary', index=False)
    
    # Scenario 2: Multi-table analysis
    print("Creating multi-table test file...")
    sales_data = generate_sales_data()
    customer_data = generate_customer_data()
    inventory_data = generate_inventory_data()
    
    with pd.ExcelWriter(output_dir / "multi_table_business.xlsx", engine='openpyxl') as writer:
        sales_data.to_excel(writer, sheet_name='Sales', index=False)
        customer_data.to_excel(writer, sheet_name='Customers', index=False)
        inventory_data.to_excel(writer, sheet_name='Inventory', index=False)
    
    # Scenario 3: Complex structure with merged cells (requires openpyxl manipulation)
    print("Creating complex structure test file...")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Complex_Report"
    
    # Create header with merged cells
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Sales Report Q4 2024'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add subheaders
    headers = ['Date', 'Product', 'Category', 'Sales', 'Notes']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add some sample data
    sample_data = [
        ['2024-12-01', 'Laptop', 'Electronics', 1500, 'Good quarter'],
        ['2024-12-02', 'Mouse', 'Accessories', 25, ''],
        ['2024-12-03', 'Monitor', 'Electronics', 300, 'High demand'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add merged cells for notes section
    ws.merge_cells('A10:E10')
    ws['A10'] = 'Additional Notes and Comments'
    ws['A10'].font = Font(bold=True)
    
    wb.save(output_dir / "complex_structure.xlsx")
    
    print(f"Test files created in: {output_dir}")
    print("Files created:")
    print("- single_table_sales.xlsx (Single table analysis)")
    print("- multi_table_business.xlsx (Multi-table analysis)")  
    print("- complex_structure.xlsx (Complex structure with merged cells)")


if __name__ == "__main__":
    create_test_excel_files()
    print("Synthetic test data generation completed!")