"""Structure Scan Agent for detecting merged cells, charts, images, and formulas."""

from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from .base import BaseAgent
from ..models.agents import StructureScanRequest, StructureScanResponse
from ..models.base import AgentRequest, AgentResponse, AgentStatus
from ..utils.config import config


class StructureScanAgent(BaseAgent):
    """Agent responsible for detecting merged cells, charts, images, and formulas."""
    
    def __init__(self):
        super().__init__(
            name="StructureScanAgent",
            description="Detects merged cells, charts, images, and formulas in Excel sheets"
        )
    
    async def process(self, request: AgentRequest) -> AgentResponse:
        """Process structure scan request."""
        if not isinstance(request, StructureScanRequest):
            return self.create_error_response(
                request,
                f"Invalid request type. Expected StructureScanRequest, got {type(request)}"
            )
        
        try:
            # Get file metadata from context
            file_metadata = request.context.get('file_metadata')
            if not file_metadata:
                return self.create_error_response(
                    request,
                    "File metadata not found in request context"
                )
            
            file_path = Path(file_metadata['file_path'])
            
            # Scan structure
            structure_info = await self._scan_sheet_structure(
                file_path, 
                request.sheet_name
            )
            
            self.logger.info(
                f"Structure scan completed for sheet '{request.sheet_name}': "
                f"{len(structure_info['merged_cells'])} merged cells, "
                f"{len(structure_info['charts'])} charts, "
                f"{len(structure_info['images'])} images, "
                f"{len(structure_info['formulas'])} formulas"
            )
            
            return StructureScanResponse(
                agent_id=self.name,
                request_id=request.request_id,
                status=AgentStatus.SUCCESS,
                merged_cells=structure_info['merged_cells'],
                charts=structure_info['charts'],
                images=structure_info['images'],
                formulas=structure_info['formulas'],
                result=structure_info
            )
            
        except Exception as e:
            self.logger.error(f"Error during structure scan: {e}")
            return self.create_error_response(request, str(e))
    
    async def _scan_sheet_structure(
        self, 
        file_path: Path, 
        sheet_name: str
    ) -> Dict[str, Any]:
        """Scan sheet structure for merged cells, charts, images, and formulas."""
        try:
            workbook = openpyxl.load_workbook(
                file_path,
                read_only=False,
                data_only=False,
                keep_links=True
            )
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            
            sheet = workbook[sheet_name]
            
            # Scan merged cells
            merged_cells = self._scan_merged_cells(sheet)
            
            # Scan charts
            charts = self._scan_charts(sheet)
            
            # Scan images
            images = self._scan_images(sheet)
            
            # Scan formulas
            formulas = self._scan_formulas(sheet)
            
            workbook.close()
            
            return {
                'merged_cells': merged_cells,
                'charts': charts,
                'images': images,
                'formulas': formulas
            }
            
        except Exception as e:
            self.logger.error(f"Error scanning sheet structure: {e}")
            raise
    
    def _scan_merged_cells(self, sheet) -> List[str]:
        """Scan for merged cells in the sheet."""
        merged_cells = []
        
        for merged_range in sheet.merged_cells.ranges:
            range_str = str(merged_range)
            merged_cells.append(range_str)
        
        return merged_cells
    
    def _scan_charts(self, sheet) -> List[Dict[str, Any]]:
        """Scan for charts in the sheet."""
        charts = []
        
        if hasattr(sheet, '_charts'):
            for i, chart in enumerate(sheet._charts):
                chart_info = {
                    'id': f"chart_{i}",
                    'type': type(chart).__name__,
                    'title': getattr(chart, 'title', None),
                    'anchor': None,
                    'width': None,
                    'height': None
                }
                
                # Get anchor information if available
                if hasattr(chart, 'anchor'):
                    anchor = chart.anchor
                    if anchor:
                        chart_info['anchor'] = {
                            'col': anchor.col,
                            'row': anchor.row,
                            'colOff': getattr(anchor, 'colOff', 0),
                            'rowOff': getattr(anchor, 'rowOff', 0)
                        }
                
                # Get dimensions if available
                if hasattr(chart, 'width'):
                    chart_info['width'] = chart.width
                if hasattr(chart, 'height'):
                    chart_info['height'] = chart.height
                
                # Get chart title if available
                if hasattr(chart, 'title') and chart.title and hasattr(chart.title, 'tx'):
                    if hasattr(chart.title.tx, 'rich'):
                        chart_info['title'] = chart.title.tx.rich.p[0].r[0].t if chart.title.tx.rich.p else None
                    elif hasattr(chart.title.tx, 'strRef'):
                        chart_info['title'] = str(chart.title.tx.strRef)
                
                charts.append(chart_info)
        
        return charts
    
    def _scan_images(self, sheet) -> List[Dict[str, Any]]:
        """Scan for images in the sheet."""
        images = []
        
        if hasattr(sheet, '_images'):
            for i, image in enumerate(sheet._images):
                image_info = {
                    'id': f"image_{i}",
                    'name': getattr(image, 'name', f"image_{i}"),
                    'format': None,
                    'anchor': None,
                    'width': None,
                    'height': None
                }
                
                # Get image format if available
                if hasattr(image, 'format'):
                    image_info['format'] = image.format
                
                # Get anchor information
                if hasattr(image, 'anchor'):
                    anchor = image.anchor
                    if anchor:
                        image_info['anchor'] = {
                            'col': getattr(anchor, 'col', None),
                            'row': getattr(anchor, 'row', None),
                            'colOff': getattr(anchor, 'colOff', 0),
                            'rowOff': getattr(anchor, 'rowOff', 0)
                        }
                
                # Get dimensions
                if hasattr(image, 'width'):
                    image_info['width'] = image.width
                if hasattr(image, 'height'):
                    image_info['height'] = image.height
                
                images.append(image_info)
        
        return images
    
    def _scan_formulas(self, sheet) -> List[str]:
        """Scan for formulas in the sheet."""
        formulas = []
        
        # Limit scan to reasonable area to avoid performance issues
        max_row = min(sheet.max_row or 1000, 1000)
        max_col = min(sheet.max_column or 100, 100)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Check if cell contains a formula
                if cell.data_type == 'f' and cell.value:
                    formula_info = f"{cell.coordinate}={cell.value}"
                    formulas.append(formula_info)
                    
                    # Limit number of formulas to avoid memory issues
                    if len(formulas) >= 500:
                        break
            
            if len(formulas) >= 500:
                break
        
        return formulas
    
    def get_detailed_merged_info(self, merged_range: str, sheet) -> Dict[str, Any]:
        """Get detailed information about a merged cell range."""
        try:
            from openpyxl.utils import range_boundaries
            
            min_col, min_row, max_col, max_row = range_boundaries(merged_range)
            
            # Get the top-left cell value
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            
            return {
                'range': merged_range,
                'top_left': f"{chr(64 + min_col)}{min_row}",
                'bottom_right': f"{chr(64 + max_col)}{max_row}",
                'value': top_left_cell.value,
                'data_type': top_left_cell.data_type,
                'rows': max_row - min_row + 1,
                'columns': max_col - min_col + 1,
                'alignment': {
                    'horizontal': top_left_cell.alignment.horizontal,
                    'vertical': top_left_cell.alignment.vertical
                } if top_left_cell.alignment else None
            }
            
        except Exception as e:
            self.logger.warning(f"Error getting detailed merged info for {merged_range}: {e}")
            return {'range': merged_range, 'error': str(e)}